from tkinter import *

#This is the master class that runs all the other functions in the program and layers them to be accessed 
class RestaurantManager(Tk):
       def __init__(self, *args, **kwargs):

              Tk.__init__(self, *args, **kwargs)
              self.frames = {}
              FunctionNames = [MainMenu,FloorPlanMenu,CustomerDetails, ReservationMenu, CreateReservation,
                               TableManager, AutomatedEmails,ExitProgram,FloorPlanCreator,CreateNewCustomer,
                               CustomerDetailsEditor,ReservationEditor,CreateEmail]

              for x in FunctionNames:
                     self.CreateFunctions(x,MainMenu)
              
       #This creates and resets the different function's screens to default when they are exited then revisitied
       def CreateFunctions(self, FunctionsToCreate, FunctionToShow):

              F = FunctionsToCreate

              container = Frame(self)
              container.grid(row=0, column=0, sticky="nsew")
              frame = F(self, container)
              self.frames[F] = frame
              frame.grid(row=0, column=0, sticky="nsew")

              self.show_frame(FunctionToShow)

              
       #This is used to move from one function to another by making it the visible layer - doesn't reset previous screen however
       def show_frame(self, cont):

              frame = self.frames[cont]
              frame.tkraise()

#This handles Main Menu Functions         
class MainMenu(Frame):

       def __init__(self, parent, controller):
              Frame.__init__(self,parent)
              Frame.config(self, bg= "grey")


              # basic reuable button/label configuration
              self.button_width = 20
              self.button_height = 3
              self.buttonbg = 'steelblue'
              self.buttonfg = 'black'
              self.buttonrelief= 'raised'

              # the information to be passed into the label - text , row , column , columnspan
              self.TitleText("Welcome to the Restaurant Management Program!", 0 ,0,3)

              #the information to be passed into the buttons-text, row, column, columnspan and the next screen to navigate to
              self.MenuButtons('Floor Plan Menu', 2, 1 , FloorPlanMenu)
              self.MenuButtons('Customer Details Menu', 3, 1 , CustomerDetails)
              self.MenuButtons('Reservations Menu', 4, 1 , ReservationMenu)
              self.MenuButtons('Table Manager Menu', 5, 1 , TableManager)
              self.MenuButtons('Automated Emails Menu', 6, 1, AutomatedEmails)
              self.MenuButtons('Exit Program', 7, 1, ExitProgram)

              self.grid_columnconfigure(1, weight=1)

       # this takes the information passed before and forms the title label introducing the program
       def TitleText(self, char_ , x_ , y_ , span):


              self.Greeting = Label(self, text = char_, width = 75, height = self.button_height,
                                    bg = self.buttonbg , fg =self.buttonfg , borderwidth = 5, relief =self.buttonrelief)
              self.Greeting.grid(row = x_, column = y_ , columnspan = span)

       
       # this takes the information passed before and uses it to create the Menu Buttons
       def MenuButtons(self, char_, x_, y_ , n):
              
              self.b = Button(self, text = char_, width = self.button_width, height = self.button_height,
                              bg = self.buttonbg, fg =self.buttonfg , pady = 10 ,command=lambda: self.master.show_frame(n))
              self.b.grid(row=x_, column=y_ , pady = (3,3))

             
#This handles Floor Plan Functions      
class FloorPlanMenu(Frame):
       def __init__(self, parent, controller):
              Frame.__init__(self, parent)
              Frame.config(self, bg= "grey")

       #This section form the Floor Plan Menu that allows access to the other sub-functions as well as a return to the Main Menu
       
              # this takes the information passed and forms the title label introducing the program
              def TitleText(char_ , x_ , y_ , span):

                     self.Greeting = Label(self, text = char_, width = 75, height = 5, bg = self.buttonbg,
                                           fg =self.buttonfg , borderwidth = 5, relief =self.buttonrelief)
                     self.Greeting.grid(row = x_, column = y_ , columnspan = span)

              
              # this takes the information passed and uses it to create the Menu Buttons
              def MenuButtons(char_, x_, y_ , n , z):
                     
                     self.b = Button(self, text = char_, width = self.button_width, height = 3, bg = self.buttonbg,
                                     fg =self.buttonfg , pady = 10 ,command=lambda: self.master.CreateFunctions(n,z))
                     self.b.grid(row=x_, column=y_ , pady = (5,5))

              # basic reuable button/label configuration
              self.button_width = 20
              self.buttonbg = 'steelblue'
              self.buttonfg = 'black'
              self.buttonrelief= 'raised'

              
              # the information to be passed into the label - text , row , column , columnspan
              TitleText("Floor Plan Management Menu \n \n Choose what you want to do!", 0 ,0,3)

              #the information to be passed into the buttons - text , row , column , columnspan and the next screen to navigate to
              MenuButtons('Create New Floor Plan', 2, 1 , FloorPlanCreator,FloorPlanCreator)
              MenuButtons('View and Edit \n Existing Floor Plans', 3, 1, FloorPlanEditor,FloorPlanEditor)
              MenuButtons('Return to Menu', 4, 1, MainMenu, MainMenu)

              self.grid_columnconfigure(1, weight=1)
       
#This handles Customer Details Functions                
class CustomerDetails(Frame):
       def __init__(self, parent, controller):
              Frame.__init__(self, parent)
              Frame.config(self, bg= "grey")

              # this takes the information passed and forms the title label introducing the program
              def TitleText(char_ , x_ , y_ , span):

                     self.Greeting = Label(self, text = char_, width = 75, height = 5, bg = self.buttonbg,
                                           fg =self.buttonfg , borderwidth = 5, relief =self.buttonrelief)
                     self.Greeting.grid(row = x_, column = y_ , columnspan = span)

              
              # this takes the information passed and uses it to create the Menu Buttons
              def MenuButtons(char_, x_, y_ , n):
                     
                     self.b = Button(self, text = char_, width = self.button_width, height = 3, bg = self.buttonbg,
                                     fg =self.buttonfg , pady = 10 ,command=lambda: self.master.show_frame(n))
                     self.b.grid(row=x_, column=y_ , pady = (5,5))


              # basic reuable button/label configuration
              self.button_width = 20
              self.buttonbg = 'steelblue'
              self.buttonfg = 'black'
              self.buttonrelief= 'raised'

              
              # the information to be passed into the label - text , row , column , columnspan
              TitleText("Customer Details Management Menu \n \n Choose what you want to do!", 0 ,0,3)


              #the information to be passed into the buttons - text , row , column , columnspan and the next screen to navigate to
              MenuButtons('Create New Customer', 2, 1 , CreateNewCustomer)
              MenuButtons('View and Edit \n Existing Customers', 3, 1, CustomerDetailsEditor)
              MenuButtons('Return to Menu', 4, 1, MainMenu)

              self.grid_columnconfigure(1, weight=1)

#This handles the Reservation Menu allowing navigation to other Reservation SubFunctions
class ReservationMenu(Frame):
       def __init__(self, parent, controller):
              Frame.__init__(self, parent)
              Frame.config(self, bg= "grey")

              #This sections form the Reservation Menu that allows access to the other sub-functions as well as a return to the Main Menu
       
              # this takes the information passed and forms the title label introducing the program
              def TitleText(char_ , x_ , y_ , span):

                     self.Greeting = Label(self, text = char_, width = 75, height = 5, bg = self.buttonbg , fg =self.buttonfg , borderwidth = 5, relief =self.buttonrelief)
                     self.Greeting.grid(row = x_, column = y_ , columnspan = span)

              
              # this takes the information passed and uses it to create the Menu Buttons
              def MenuButtons(char_, x_, y_ , n):
                     
                     self.b = Button(self, text = char_, width = self.button_width, height = 3, bg = self.buttonbg , fg =self.buttonfg , pady = 10 ,command=lambda: self.master.show_frame(n))
                     self.b.grid(row=x_, column=y_ , pady = (5,5))


              # basic reuable button/label configuration
              self.button_width = 20
              self.buttonbg = 'steelblue'
              self.buttonfg = 'black'
              self.buttonrelief= 'raised'

              
              # the information to be passed into the label - text , row , column , columnspan
              TitleText("Reservation Management Menu \n \n Choose what you want to do!", 0 ,0,3)


              #the information to be passed into the buttons - text , row , column , columnspan and the next screen to navigate to
              MenuButtons('Create New Reservation', 2, 1 , CreateReservation)
              MenuButtons('View and Edit \n Existing Reservations', 3, 1, ReservationEditor)
              MenuButtons('Return to Menu', 4, 1, MainMenu)

              self.grid_columnconfigure(1, weight=1)

#This handles the Reservation Creation SubFunction
class CreateReservation(Frame):
       def __init__(self, parent, controller):
              Frame.__init__(self, parent)
              Frame.config(self, bg= "grey")
              
              self.grid_columnconfigure(1, weight=1)

              #List of titles to be used when creating labels for the entry boxes
              infolabels = [
                     "Customer Name :" ,
                     "Number of Visitors :" ,
                     "Date of Visit :" ,
                     "Time of Visit :",
                     "Choose Floor Plan to Apply to :"
                     ]

              #List of titles to be put inside the entry boxes
              EntryBoxText = [
                     "Enter Customer name here" ,
                     "Enter number of Visitors (Max = 20)" ,
                     "Date - dd/mm/yy " ,
                     "Time (Between 9,00 and 21,00) - hh,mm"
                     ]

              #Creates a title telling the user what to do
              TitleText = Label(self, text="Please enter the reservation details into the correct input boxes",  bg = 'steelblue',
                                fg = 'black' , borderwidth = 5, pady = 10, padx=3 , relief= 'raised', font=("Courier New", 10))
              TitleText.grid(row = 0 , column = 0 , columnspan = 3)

              #Uses a for loop to create the information labels next to each entry box
              for x in range (0,5):
                     InfoLabel = Label(self, text= infolabels[x], bg = "grey", fg = 'black' ,
                                       borderwidth = 5, pady = 15 , padx = 10, font=("Courier New", 10))
                     InfoLabel.grid(row = x + 1 , column = 0 , sticky = W)

              #Uses a for loop to create each entry box
              self.inputDict = {}
              for x in range (0,4):
                     self.inputDict[x] = Entry(self , width = 52, fg = 'white' , bg = 'black' , borderwidth = 10 )
                     self.inputDict[x].grid(row= x+1, column = 1, columnspan = 2)
                     self.inputDict[x].insert(0, EntryBoxText[x])#Enters the entry box text 

              #Function to let the user choose the floor plan and reserve tables on it
              ChooseFloorPlanBtn = Button(self, text = "Choose the floor plan you want \n the reservation to be applied to", command = self.ChoosePlan,
                                          bg = 'steelblue'  , fg = 'black' , pady = 10, padx = 10 , relief= 'raised', font=("Courier New", 10))
              ChooseFloorPlanBtn.grid(row = 5, column = 1, pady = (5,20))

              #Uses entered data to find an appropriate reservation
              SubmitR = Button(self,  text = "Search for Available Reservation Slots ",  bg = 'steelblue', command = self.ReservationSearch,
                               fg = 'black' , pady = 10, padx = 10 , relief= 'raised', font=("Courier New", 10))
              SubmitR.grid(row = 6, column = 1)

              #Opens an exit box allowing the user to choose how to exit the subfunction
              Exit = Button(self,  text = "Return to Menu ",  bg = 'steelblue'  , fg = 'black' , pady = 10 , relief= 'raised',
                            font=("Courier New", 10) , command = self.ReservExitWarning)
              Exit.grid(row = 6, column = 0)

       def ChoosePlan(self):
              #Activates when a table has ben clicked to be reserved
              def tableReserver (n):
                     self.seatchoice = Toplevel(width=500, height =800 , bg='grey')#Opens a new window to house the GUI
                     #Label to ask the user whether they are sure they want to reserve this seat
                     chooselabel = Label(self.seatchoice,text = "Are you sure you want to reserve this seat? ",font = "Verdana 20",
                                         pady = 10,padx = 20,  bg = 'steelblue'  , fg = 'black' , relief = "raised" , borderwidth = 5)
                     chooselabel.grid(row=0,column=0,columnspan=2)
                     #Yes button that will save the table name by activating the selected function
                     reserve = Button (self.seatchoice, text = "Yes ", font = "Verdana 15" , bg = "steelblue",
                                          pady = 10,command=lambda n=n:selected(n , "purple"))
                     reserve.grid(row=1,column=0)
                     #No Button to close th function and allow another tabel to be chosen
                     noReserve = Button (self.seatchoice, text = "No ", font = "Verdana 15" , bg = "steelblue",
                                          pady = 10,command= self.seatchoice.destroy)
                     noReserve.grid(row=1,column=1)

              #When a user has selected a table to reserve, this function changs it colour to purple and saves the table number
              def selected (n , x):
                     seatingcolourdict[n] = x
                     tabledict[n].configure(bg = x)
                     self.seatchoice.destroy()
                  
              #Function that activates when save to file button pressed - iterates through the tables until it finds a reserved one
              #It then savs the table numbr to the dictionary containing all the reservation details
              def saveandexit():
                     count = 0
                     for z in seatingcolourdict :
                            if seatingcolourdict[z] == "purple":
                                   self.inputDict[5+count] = ("Table " + str(z))
                                   count +=1
                     TableReserv.destroy()   

              seatingcolourdict = {} #Dictionary holding table and corresponding colour
              tabledict = {} #Dictionary holding all the tables and their functions/information

              import glob

              if len(glob.glob("*.txt")) == 0: #Checks whether there aree any .txt floor plans created

                     def returntofloormenu(): #Returns the user to main menu
                            self.master.CreateFunctions(MainMenu,MainMenu)
                            noPlans.destroy()
                            
                     #If no floor plans, it opens a new window to inform the user
                     noPlans = Toplevel(width=800, height =800 , bg='grey')
                     noFloors = Label(noPlans, text = "No saved floor plans. Create one, then return if needed.",
                                         bg = 'steelblue',width = 70, font=("Courier New", 10))
                     noFloors.grid(row = 0, column = 0 , columnspan = 4)

                     returntomenu = Button (noPlans , text = "Return to Menu" , command= returntofloormenu, bg = 'steelblue')
                     returntomenu.grid(row=2,column=0,columnspan=4)

              else:
                      
                     from tkinter.filedialog import askopenfilename #Allows us to use tkinter file management
                     # show an "Open" dialog box and return the path to the selected file
                     filename = askopenfilename(filetypes = [("Text Files", ".txt")])

                     import os
                     self.inputDict[4] = os.path.basename(filename)
                     
                     #Opens the chosen floor plan file and reads the colour of each table and saves it to a list
                     textfile = open(filename,"r") 
                     with open(filename) as f:
                         seatingcolourlist = f.read().splitlines()
                     
                     count1 = 1 # This iterates through the list and converts it to a dictionary
                     for x in seatingcolourlist :
                            seatingcolourdict[str(count1)] = x
                            count1+=1

                     TableReserv = Toplevel(width=600, height =800 , bg='grey')

                     info = Label (TableReserv, text = "Click a table to reserve it ", font = "Verdana 15", padx = 10, pady = 10,
                                   bg = 'steelblue'  , fg = 'black' , relief = "raised" , borderwidth = 5)
                     info.grid(row=0,column=1, columnspan = 3) # Creates a label to inform the user

                     twoseater = Label (TableReserv, text = "2 Seats ", font = "Verdana 15" , bg = "yellow", padx = 5).grid(row=1,column=0)
                     fourseater = Label (TableReserv, text = "4 Seats  ", font = "Verdana 15" , bg = "orange", padx = 5).grid(row=1,column=1)
                     sixseater = Label (TableReserv, text = "6 Seats ", font = "Verdana 15" , bg = "red", padx = 5).grid(row=1,column=2)
                     tenseater = Label (TableReserv, text = "10 Seats ", font = "Verdana 15" , bg = "cyan", padx = 5).grid(row=1,column=3)
                     reserved = Label (TableReserv, text = "Reserved ", font = "Verdana 15" , bg = "purple", padx = 5).grid(row=1,column=4)
              
                     rowNumber = 2
                     columnNumber = 0

                     colourvalues = {"yellow" : "2" , "orange" : "4" , "red" : "6" , "cyan" : "10"}
                     #Creates the grid of tables, coloured according to the saved floorplan
                     for n in seatingcolourdict:
                            tabledict[n] = Button(TableReserv,text = "Table " + str(n), bg = seatingcolourdict[n],
                                                  font = "Verdana 12" ,width = 10, height = 5,command=lambda n=n: tableReserver(n))
                            tabledict[n].grid(row = rowNumber, column = columnNumber, padx = 5, pady = 10)

                            columnNumber += 1
                            if columnNumber == 5:
                                   columnNumber = 0
                                   rowNumber += 1

                     #Button at bottom of screen to run nameenter function and then save and exit
                     saveandexitbtn = Button (TableReserv, text = "Save Floor Plan and Exit to Menu" , pady = 10 , padx = 10, font = "Verdana 12",
                                              command =saveandexit ,  bg = 'steelblue'  , fg = 'black' , relief = "raised" , borderwidth = 5 )
                     saveandexitbtn.grid(row = 7 , column = 1 , columnspan = 3)
                     #Small fix to keep GUI looking clean
                     self.grid_columnconfigure(1, weight=1)
                     self.grid_columnconfigure(3, weight=1)
              
       

       def ReservationSearch(self):
              import datetime #Imports the datetime modle to help validate entries

              #Function that creates a new window with an error message if any entry is incorrect
              def ErrorMessage(): 
                     errorMessage = Toplevel(bg = "gray")

                     #A label warning the the user data may be lost
                     msg = Label(errorMessage, bg = 'steelblue'  , fg = 'black', relief= 'raised', borderwidth = 5, pady = 15 , padx = 10,
                                 text="Look carefully at the Inputted Data. Make sure :\n The name contains no Numbers,\n The visitor size is less than 21,\n The entered time and date are valid,\n and a table has ben reserved.",
                                 )
                     msg.pack()

              def ChooseAnother():
                     booked= Toplevel(bg = "gray")

                     #A label warning the the user data may be lost
                     msg = Label(booked, text="Your specific reservation is already booked. Change tables, time or date and try again.",
                                 bg = 'steelblue'  , fg = 'black', relief= 'raised', borderwidth = 5, pady = 15 , padx = 10,)
                     msg.pack()
                     
              def SaveBooking():

                     import openpyxl #Excel file manager module
                     import datetime

                     #Opens excel file and sheet containg database
                     theFile = openpyxl.load_workbook('ReservationForm.xlsx')
                     currentSheet = theFile['Sheet1']

                     Y = currentSheet.max_row+1 #Selects the next empty row
                     X=0

                     #Inputs opt in or out into database
                     currentSheet.cell(column= 1, row=Y, value= self.inputtedinfo[0])
                     currentSheet.cell(column= 2, row=Y, value= int(self.inputtedinfo[1]))

                     reservDate = datetime.datetime.strptime(self.inputtedinfo[2], "%d/%m/%y")
                     reservDate = reservDate.strftime("%d/%m/%y")
                     currentSheet.cell(column= 3, row=Y).value= reservDate
                     
                     currentSheet.cell(column= 4, row=Y, value= self.inputtedinfo[3])
                     currentSheet.cell(column= 5, row=Y, value= self.inputDict[4])
                     currentSheet.cell(column= 6, row=Y, value= self.inputDict[5])
                     currentSheet.cell(column= 7, row=Y, value= self.inputDict[6])

                     #Saves the excel file               
                     theFile.save('ReservationForm.xlsx')
              
              #Function used to check if a time entered is in the restaurants iopening hours
              def time_in_range(start, end, x):                  
                  if start <= end:
                      return start <= x <= end
                  else:
                      return start <= x or x <= end

              #Function to check the entries to see if they are valid            
              def Validation():

                     #Creates a list of the entered values
                     self.inputtedinfo.clear()
                     for x in range(0,4):

                            info = self.inputDict[x].get()
                            self.inputtedinfo.append(info)

                     input_string = self.inputtedinfo[3] #Splits up the time format so checking works
                     self.output = input_string.split(',')

                     count = 0
                     
                     if any(char.isdigit() for char in self.inputtedinfo[0]): # Checks if name contains a number
                            ErrorMessage()#If it does it pulls up error message
                            count +=1

                     elif int(self.inputtedinfo[1]) > 20 : # Checks if the enters amount of guests is more than the can be split amongst 2 largest tables
                            ErrorMessage()#If there are too many customers it pulls up error message
                            count +=1

                     elif int(self.inputtedinfo[1]) < 1 : # Checks if the enters amount of guests is less than 0
                            ErrorMessage()#If there are too few customers it pulls up error message
                            count +=1

                     elif time_in_range(start, end, datetime.time(int(self.output[0]),int(self.output[1]))) == False:# Checks if time entered is during operational hours
                            ErrorMessage()#If not it it pulls up error message
                            count +=1

                     #Used to check if a floor plan was selected 
                     try :
                            print(self.inputDict[4])
                            
                     except:
                            ErrorMessage()#If not it it pulls up error message
                            count +=1

                     #Used to check if at least one table was reserved
                     try :
                            print(self.inputDict[5])
                            
                     except:
                            ErrorMessage()#If not it it pulls up error message
                            count +=1
                            
                     #Used to check if the entered date is valid                     
                     try:
                            date_of_birth = datetime.datetime.strptime(self.inputtedinfo[2], "%d/%m/%y")
                     except:
                            ErrorMessage()#If not valid it pulls up error message
                            count +=1

                     if count == 0:
                            Search()

              def Search():
                     count = 0
                     import openpyxl #Excel file manager module
                     import datetime
                     #Opens excel file and sheet containg database and selects the fist column (customer names)
                     theFile = openpyxl.load_workbook('ReservationForm.xlsx')
                     currentSheet = theFile['Sheet1']

                     #Iterates through the date column to find if their is a reservation on the same date
                     print(currentSheet.max_row)
                     for i in range(2,currentSheet.max_row+1):
                            if currentSheet.cell(row=i, column=3).value.strftime("%d/%m/%y") == self.inputtedinfo[2]:
                                   if currentSheet.cell(row=i, column=4).value == self.inputtedinfo[3]:#If same date, check if same time
                                          if currentSheet.cell(row=i, column=5).value == self.inputDict[4]:#If same time check if same floorplan
                                                 #If same floorplan check if any tables overlap
                                                 if currentSheet.cell(row=i, column=6).value == self.inputDict[5 or 6]:
                                                        ChooseAnother()
                                                        count+=1

                                                 if currentSheet.cell(row=i, column=7).value == self.inputDict[5 or 6]:
                                                        ChooseAnother()
                                                        count+=1
                     if count == 0:
                            SaveBooking()

              self.inputtedinfo = [] #List of entered values
              #Restaurant opening/closing times
              start = datetime.time(9, 0)
              end = datetime.time(21, 0)
              #Runs the validation
              Validation()
                            
                            
              
       #An exit box containing choices to leave or stay in reservation creation
       def ReservExitWarning(self):

              #Creates a top level window to hold the exit box
              Warning = Toplevel(bg = "gray")

              #A label warning the the user data may be lost
              msg = Label(Warning, text="Are you sure you want to return to Main Menu? Current entries may be deleted or saved for later.",
                          bg = 'steelblue'  , fg = 'black', relief= 'raised', borderwidth = 5, pady = 15 , padx = 10,)
              msg.pack()

              #Closes the exit box
              def ReservReturn():

                     Warning.destroy()

              #Uses the master function to reset the reservation creation function and return to menu
              def MenuReturn():

                     self.master.CreateFunctions(CreateReservation,ReservationMenu)
                     Warning.destroy()
              
              #Returns user to menu but keeps entered data in box to be returned to
              def TempMenuReturn():

                     self.master.show_frame(MainMenu)
                     Warning.destroy()

              #Creates the button to close the exit box
              ReturnToReservBtn = Button(Warning, text="Return to Reservation Creation", command= ReservReturn, bg = 'steelblue'  , fg = 'black', relief= 'raised', borderwidth = 5, pady = 5 , padx = 10,)
              ReturnToReservBtn.pack()

              #Creates the button to return to menu and clear entries
              ReturnToMenuBtn = Button(Warning, text="Return to Menu and Clear Entries", command= MenuReturn, bg = 'steelblue'  , fg = 'black', relief= 'raised', borderwidth = 5, pady = 5 , padx = 10,)
              ReturnToMenuBtn.pack()

              #Creates the button to return to menu but keep entries
              TempReturnToMenuBtn = Button(Warning, text="Return to Menu But Keep Entries Saved", command= TempMenuReturn, bg = 'steelblue'  , fg = 'black', relief= 'raised', borderwidth = 5, pady = 5 , padx = 10,)
              TempReturnToMenuBtn.pack()


#This handles the Reservation Creation SubFunction
class ReservationEditor(Frame):
       def __init__(self, parent, controller):
              Frame.__init__(self, parent)
              Frame.config(self, bg= "grey")
              
              self.grid_columnconfigure(1, weight=1)

              self.Inputdict = {}
              self.variabledict = {}

              CUSTOMERS = []

              def InputCustomer(customerName):
                     import datetime
                     #Stores a selected reservation's details to be inputted into the form
                     self.selectedReservation = []                     

                     import openpyxl #Excel file manager module
                     #Opens excel file and sheet containg database and selects the fist column (customer names)
                     theFile = openpyxl.load_workbook('ReservationForm.xlsx')
                     currentSheet = theFile['Sheet1']

                     #Iterates through the name column to find the row containing the selected reservation
                     for i in range(1,currentSheet.max_row+1):
                            if currentSheet.cell(row=i, column=1).value == customerName:
                                   self.customerdetailrow = i
                                   #Inputs selected custoemrs setails into the list
                                   for j in range(1, currentSheet.max_column+1):
                                          self.selectedReservation.append(currentSheet.cell(row=i, column=j).value)

                     #Converts from datetime to just date
                     reservDate = self.selectedReservation[2]
                     reservDate = reservDate.strftime("%d/%m/%y")
                     self.selectedReservation[2] = reservDate

                     #Iterates through the entry fields and inputs the correct data from the selectedcustomer list
                     for x in range (1,8):
                            self.reservationdetailsdict[x].delete(0, END)
                            self.reservationdetailsdict[x].insert(0,self.selectedReservation[x-1])


              #Activated when save button pressed. Opens the reservation database and enters information
              def savedetails():
                     import openpyxl #Excel file manager module

                     #Opens excel file and sheet containg database
                     theFile = openpyxl.load_workbook('ReservationForm.xlsx')
                     currentSheet = theFile['Sheet1']

                     X = 1
                     Y = self.customerdetailrow #Selects the row containing the reservation to edit
                     
                     #Inputs info into database
                     for n in self.reservationdetailsdict:
                            enteredvalue = self.reservationdetailsdict[n].get()
                            print(enteredvalue)

                            currentSheet.cell(column= X, row=Y, value= enteredvalue)
                            X += 1

                     #Saves the excel file               
                     theFile.save('ReservationForm.xlsx')


              #Creates a title label to tell the user what to do
              TitleText = Label(self, text="Please select a reservation. Its details can then be edited.",  bg = 'steelblue'
                                ,fg = 'black' , borderwidth = 5, pady = 10 , padx = 10, relief= 'raised', font=("Courier New", 10))
              TitleText.grid(row = 0 , column = 0 , columnspan = 3, pady = (5,10))


              import openpyxl #Excel file manager module

              #Opens excel file and sheet containg database and selects the fist column (customer names)
              theFile = openpyxl.load_workbook('ReservationForm.xlsx')
              currentSheet = theFile['Sheet1']
              first_column = currentSheet['A']

              #Program iterates through names and adds them to a list to be accessed by the menu
              for x in range(len(first_column)):
                     if currentSheet.cell(row=x+1, column=1).value == None:
                            pass
                     else:
                            CUSTOMERS.append(first_column[x].value) 

              #Removes the header of the column (Names) from the choices
              CUSTOMERS.remove("Name")

              #Creates the options menu with all customer names in the list
              variable = StringVar()
              SelectCustomer = OptionMenu(self, variable, *CUSTOMERS, command = InputCustomer)
              SelectCustomer.grid(row = 1 , column = 1 , columnspan = 2, pady = (5,10))

              LABELS = ["Name  :", "Number of Visitors  :" , "Date of Visit  :",
                        "Time of Visit :", "Floor Plan :", "Table Choice(1) :", "Table Choice(2) :",]

              ReservInfo = {}

              for x in range(0,7):
                     #Creates the labels for each entry field
                     ReservInfo[x] = Label(self, text = LABELS[x], bg = 'gray')
                     ReservInfo[x].grid(row = 2 + x, column = 0, sticky = W , pady = (5,5))
              
              #creates the entry fields for the details like before
              self.reservationdetailsdict = {}

              for x in range (1,8):
                     self.reservationdetailsdict[x] = Entry(self , width = 50, fg = 'white' , bg = 'black' , borderwidth = 10 )
                     self.reservationdetailsdict[x].grid(row= x+1, column = 1, columnspan = 2, pady = (5,5))

              #Save customer details button
              changeReserv = Button(self, text = "Change Floor Plan and/or Tables", bg = 'steelblue' , fg = 'black', borderwidth = 5,
                                    pady = 10 , padx = 10, relief= 'raised' ,font=("Courier New", 10), command = self.ChoosePlan )
              changeReserv.grid(row = 14, column = 1, pady = (10,10))   
              
              #Save customer details button
              savecustomer = Button(self, text = "Save customer details", bg = 'steelblue' , fg = 'black', borderwidth = 5,
                                    pady = 10 , padx = 10, relief= 'raised' ,font=("Courier New", 10), command = savedetails)
              savecustomer.grid(row = 15, column = 1, sticky = W)

              #A exit button to leave to menu
              exit = Button(self, text = "Exit Button",  bg = 'steelblue'  , fg = 'black',borderwidth = 5,
                            pady = 10 , padx = 10, relief= 'raised', font=("Courier New", 10))
              exit.grid(row = 15, column = 1, sticky = S , padx = (300,5))


       def ChoosePlan(self):

              def tableReserver (n):
                     self.seatchoice = Toplevel(width=500, height =800 , bg='grey')#Opens a new window to house the GUI
                     #Label to ask the user whether they are sure they want to reserve this seat
                     chooselabel = Label(self.seatchoice,text = "Are you sure you want to reserve this seat? ",font = "Verdana 20",
                                                pady = 10,padx = 20,  bg = 'steelblue'  , fg = 'black' , relief = "raised" , borderwidth = 5)
                     chooselabel.grid(row=0,column=0,columnspan=2)
                            #Yes button that will save the table name by activating the selected function
                     reserve = Button (self.seatchoice, text = "Yes ", font = "Verdana 15" , bg = "steelblue",
                                                 pady = 10,command=lambda n=n:selected(n , "purple"))
                     reserve.grid(row=1,column=0)
                            #No Button to close th function and allow another tabel to be chosen
                     noReserve = Button (self.seatchoice, text = "No ", font = "Verdana 15" , bg = "steelblue",
                                                 pady = 10,command= self.seatchoice.destroy)
                     noReserve.grid(row=1,column=1)

              #When a user has selected a table to reserve, this function changes it colour to purple and puts the table number into the entry field
              def selected (n , x):
                     self.reservationdetailsdict[6 + self.counting].delete(0, END)
                     self.reservationdetailsdict[6 + self.counting].insert(0, "Table " + str(n))
                     tabledict[n].configure(bg = x)
                     self.counting += 1
                     self.seatchoice.destroy()
                  

              seatingcolourdict = {} #Dictionary holding table and corresponding colour
              tabledict = {} #Dictionary holding all the tables and their functions/information
              self.counting = 0

              import glob

              if len(glob.glob("*.txt")) == 0: #Checks whether there are any .txt floor plans created

                     def returntofloormenu(): #Returns the user to main menu
                            self.master.CreateFunctions(MainMenu,MainMenu)
                            noPlans.destroy()
                            
                     #If no floor plans, it opens a new window to inform the user
                     noPlans = Toplevel(width=800, height =800 , bg='grey')
                     noFloors = Label(noPlans, text = "No saved floor plans. Create one, then return if needed.",
                                         bg = 'steelblue',width = 70, font=("Courier New", 10))
                     noFloors.grid(row = 0, column = 0 , columnspan = 4)

                     returntomenu = Button (noPlans , text = "Return to Menu" , command= returntofloormenu, bg = 'steelblue')
                     returntomenu.grid(row=2,column=0,columnspan=4)

              else:
                      
                     from tkinter.filedialog import askopenfilename #Allows us to use tkinter file management
                     # show an "Open" dialog box and return the path to the selected file
                     filename = askopenfilename(filetypes = [("Text Files", ".txt")])

                     import os
                     #Changes floor plan entry if a different one was selcted
                     self.reservationdetailsdict[5].delete(0,END)
                     self.reservationdetailsdict[5].insert(0, os.path.basename(filename))
                     
                     #Opens the chosen floor plan file and reads the colour of each table and saves it to a list
                     textfile = open(filename,"r") 
                     with open(filename) as f:
                         seatingcolourlist = f.read().splitlines()
                     
                     count1 = 1 # This iterates through the list and converts it to a dictionary
                     for x in seatingcolourlist :
                            seatingcolourdict[str(count1)] = x
                            count1+=1

                     TableReserv = Toplevel(width=600, height =800 , bg='grey') #Opens new toplevel GUI to hold the table layout
                     

                     info = Label (TableReserv, text = "Click a table to reserve it ", font = "Verdana 15", padx = 10, pady = 10,
                                   bg = 'steelblue'  , fg = 'black' , relief = "raised" , borderwidth = 5)
                     info.grid(row=0,column=1, columnspan = 3) # Creates a label to inform the user

                     twoseater = Label (TableReserv, text = "2 Seats ", font = "Verdana 15" , bg = "yellow", padx = 5).grid(row=1,column=0)
                     fourseater = Label (TableReserv, text = "4 Seats  ", font = "Verdana 15" , bg = "orange", padx = 5).grid(row=1,column=1)
                     sixseater = Label (TableReserv, text = "6 Seats ", font = "Verdana 15" , bg = "red", padx = 5).grid(row=1,column=2)
                     tenseater = Label (TableReserv, text = "10 Seats ", font = "Verdana 15" , bg = "cyan", padx = 5).grid(row=1,column=3)
                     reserved = Label (TableReserv, text = "Reserved ", font = "Verdana 15" , bg = "purple", padx = 5).grid(row=1,column=4)
              
                     rowNumber = 2
                     columnNumber = 0

                     colourvalues = {"yellow" : "2" , "orange" : "4" , "red" : "6" , "cyan" : "10"}
                     #Creates the grid of tables, coloured according to the saved floorplan
                     for n in seatingcolourdict:
                            tabledict[n] = Button(TableReserv,text = "Table " + str(n), bg = seatingcolourdict[n],
                                                  font = "Verdana 12" ,width = 10, height = 5,command=lambda n=n: tableReserver(n))
                            tabledict[n].grid(row = rowNumber, column = columnNumber, padx = 5, pady = 10)

                            columnNumber += 1
                            if columnNumber == 5:
                                   columnNumber = 0
                                   rowNumber += 1

                     #Small fix to keep GUI looking clean
                     self.grid_columnconfigure(1, weight=1)
                     self.grid_columnconfigure(3, weight=1)
              


#This handles Table Management Functions              
class TableManager(Frame):
       def __init__(self, parent, controller):
              Frame.__init__(self, parent)
              Frame.config(self, bg= "grey")

              # this takes the information passed and forms the title label introducing the program
              def TitleText(char_ , x_ , y_ , span):

                     self.Greeting = Label(self, text = char_, width = 75, height = 5, bg = self.buttonbg , fg =self.buttonfg , borderwidth = 5, relief =self.buttonrelief)
                     self.Greeting.grid(row = x_, column = y_ , columnspan = span)

              
              # this takes the information passed and uses it to create the Menu Buttons
              def MenuButtons(char_, x_, y_ , n,z):
                     
                     self.b = Button(self, text = char_, width = self.button_width, height = 3, bg = self.buttonbg , fg =self.buttonfg , pady = 10 ,command=lambda: self.master.CreateFunctions(n,z))
                     self.b.grid(row=x_, column=y_ , pady = (5,5))


              # basic reuable button/label configuration
              self.button_width = 20
              self.buttonbg = 'steelblue'
              self.buttonfg = 'black'
              self.buttonrelief= 'raised'

              
              # the information to be passed into the label - text , row , column , columnspan
              TitleText("Table Management Menu \n \n Choose what you want to do!", 0 ,0,3)


              #the information to be passed into the buttons - text , row , column , columnspan and the next screen to navigate to
              MenuButtons('Choose Floor Plan to Use', 2, 1, TableManaging, TableManaging)
              MenuButtons('Access Last Used \n Floor Plan', 3, 1 , LastTableManaging,LastTableManaging)
              MenuButtons('Return to Menu', 4, 1, MainMenu,MainMenu)

              self.grid_columnconfigure(1, weight=1)


class TableManaging(Frame):
       def __init__(self, parent, controller):
              Frame.__init__(self, parent)
              Frame.config(self, bg= "grey")

              def MenuReturn():
                     self.master.show_frame(MainMenu)

              def Refresh():

                     reservedTables = []

                     import openpyxl
                     import datetime

                     date_object = datetime.date.today()
                     date_object = date_object.strftime("%d/%m/%y")

                     theFile = openpyxl.load_workbook('ReservationForm.xlsx')
                     currentSheet = theFile['Sheet1']
                     first_column = currentSheet['A']

                     for x in range(2,len(first_column)+1):

                            floorPlan = currentSheet.cell(row=x, column=5).value

                            Date = currentSheet.cell(row=x, column=3).value
                            Date = Date.strftime("%d/%m/%y")

                            now = datetime.datetime.now()
                            now = now.strftime('%H:%M')

                            timein30 = datetime.datetime.now() + datetime.timedelta(minutes=30)
                            timein30 = timein30.strftime('%H:%M')

                            reservationTime = currentSheet.cell(row=x, column=4).value

                            date_time_obj = datetime.datetime.strptime(reservationTime,'%H,%M').time()
                            date_time_obj = date_time_obj.strftime('%H:%M')

                            if floorPlan == os.path.basename(self.filename):
                                   if Date == date_object:
                                          if now <= date_time_obj <= timein30:
                                                 reservedTables.append(currentSheet.cell(row=x, column=6).value)
                                                 reservedTables.append(currentSheet.cell(row=x, column=7).value)
                                                 CustomerRow = x

                                                 
                     for x in range(0,len(reservedTables)):
                            reservedTables[x] = int(''.join(list(filter(str.isdigit, reservedTables[x]))))


                     for x in reservedTables:
                            tabledict[str(x)].config(bg = "purple")

                            tableReservedCustomer[x] = CustomerRow
                            print (tableReservedCustomer)
                                   


              def SeatReserved(n):
                            
                     customer = (tableReservedCustomer[int(n)])

                     for x in tableReservedCustomer:
                            if tableReservedCustomer[int(x)] == customer:
                                   tabledict[str(x)].config(bg = "red")
                     
                     

              def CheckReservation(n):
                     import openpyxl #Excel file manager module

                     SendTo = []#A list of everyone to send the email to

                     #Opens excel file and sheet containg database and selects the fist column (customer names)
                     theFile = openpyxl.load_workbook('ReservationForm.xlsx')
                     currentSheet = theFile['Sheet1']
                     table1check = currentSheet['F']
                     table2check = currentSheet['G']

                     tablename = ("Table " + str(n))

                     reservationDates = {}
                     reservationTimes = {}
                     reservationCustomer = {}

                     for x in range(0,len(table1check)):
                                   if currentSheet.cell(row=x+1, column=6).value == tablename:
                                          
                                          if currentSheet.cell(row=x+1, column=5).value == os.path.basename(self.filename):

                                                 reservationCustomer[x+1] = currentSheet.cell(row=x+1, column=1).value

                                                 reservDate = currentSheet.cell(row=x+1, column=3).value
                                                 reservDate = reservDate.strftime("%d/%m/%y")
                                                 #reservationDates.append(reservDate)

                                                 reservationDates[x+1] = reservDate

                                                 reservationTimes[x+1] = currentSheet.cell(row=x+1, column=4).value

                                   elif currentSheet.cell(row=x+1, column=7).value == tablename:
                                          if currentSheet.cell(row=x+1, column=5).value == self.filename:

                                                 reservationCustomer[x+1] = currentSheet.cell(row=x+1, column=1).value

                                                 reservDate = currentSheet.cell(row=x+1, column=3).value
                                                 reservDate = reservDate.strftime("%d/%m/%y")
                                                 #reservationDates.append(reservDate)

                                                 reservationDates[x+1] = reservDate

                                                 reservationTimes[x+1] = currentSheet.cell(row=x+1, column=4).value

                     if not reservationDates:
                            
                            reservationCustomer[0] =("No reservations")
                            reservationDates[0] =("No reservations")
                            reservationTimes[0] =("No reservations")

                     samedayrows = []
                     samedaytimes = []

                     sortedReservationRows = sorted(reservationDates, key=reservationDates.get)

                     
                     if reservationDates[sortedReservationRows[0]]== reservationDates[sortedReservationRows[1]]:
                            print("Multiple next dates")

                            for x in range(len(reservationDates)):
                                   if reservationDates[sortedReservationRows[x]] == reservationDates[sortedReservationRows[0]]:
                                          samedayrows.append(sortedReservationRows[x])

                            for x in samedayrows :
                                   samedaytimes.append(currentSheet.cell(row=x, column=4).value)
                                   print(samedaytimes)

                            samedaytimes.sort()
                            print(samedaytimes)
                                   
                     else:
                            print("Single Date")

                     nextReservDate = reservationDates[sortedReservationRows[0]]
                     nextReservTime = samedaytimes[0]

                     nextReservCustomer = ' '
                     NextReservRow = 0

                     for x in range(2,len(table1check)):
                            self.ReservedDate = currentSheet.cell(row=x, column=3).value
                            self.ReservedDate = self.ReservedDate.strftime("%d/%m/%y")
                            
                            if self.ReservedDate == nextReservDate:
                                   if currentSheet.cell(row=x, column=4).value == nextReservTime:
                                          NextReservRow = x
                                          
                     
                     nextReservdetails = []
                     
                     for j in range(1, currentSheet.max_column+1):
                                   nextReservdetails.append(currentSheet.cell(row=NextReservRow, column=j).value)

                     print(nextReservdetails)

                            
                     nextReservation = Toplevel(width=500, height =800 , bg='grey')

                     nextReserv = Label(nextReservation, text = "Table's Next Reservation Information", pady = 10, width = 40,
                                       padx = 20,  bg = 'steelblue'  , fg = 'black' , relief = "raised" , borderwidth = 5)
                     nextReserv.grid(row = 0, column = 0, columnspan = 2)

                     nextReservCustomer = Label(nextReservation, text = "Table's Next Reservation Customer : ", pady = 10, width = 20,
                                       padx = 20,  bg = 'grey'  , fg = 'black' , borderwidth = 5).grid(row = 1, column = 0,padx=(5,5))

                     nextReservCustomerInfo = Label(nextReservation, text = str(nextReservdetails[0]), pady = 10, width = 20,
                                       padx = 20,  bg = 'grey'  , fg = 'black' , borderwidth = 5).grid(row = 1, column = 1)

                     nextReservDate = Label(nextReservation, text = "Table's Next Reservation Date : ", pady = 10, width = 20,
                                       padx = 20,  bg = 'grey'  , fg = 'black' , borderwidth = 5).grid(row = 2, column = 0)

                     nextReservDateInfo = Label(nextReservation, text = str(self.ReservedDate), pady = 10, width = 20,
                                       padx = 20,  bg = 'grey'  , fg = 'black' , borderwidth = 5).grid(row = 2, column = 1)

                     nextReservTime = Label(nextReservation, text = "Table's Next Reservation Time : ", pady = 10, width = 20,
                                       padx = 20,  bg = 'grey'  , fg = 'black' , borderwidth = 5).grid(row = 3, column = 0)

                     nextReservTimeInfo = Label(nextReservation, text = str(nextReservdetails[3]), pady = 10, width = 20,
                                       padx = 20,  bg = 'grey'  , fg = 'black' , borderwidth = 5).grid(row = 3, column = 1)


              def UnseatVisitor(n):
                     tabledict[n].configure(bg = "green")
                     self.SeatVistiorBtn.config(text = "Seat Visitor on Table", command=lambda n=n: SeatVisitor(n))
                     
                     try:
                            del tableSeatedCustomer[n]
                     except:
                            pass

              def ChosenCustomer(chosen):

                     customerdetails = []

                     import openpyxl

                     theFile = openpyxl.load_workbook('Customer Details Form.xlsx')
                     currentSheet = theFile['Sheet1']
                     first_column = currentSheet['A']

                     #Program iterates through names and adds them to a list to be accessed by the menu
                     for x in range(len(first_column)):
                            if currentSheet.cell(row=x+1, column=1).value == chosen:
                                   customerrow= x+1
                                   #Inputs selected custoemrs setails into the self.selectedcustomer list
                                   for j in range(1, currentSheet.max_column+1):
                                          customerdetails.append(currentSheet.cell(row=customerrow, column=j).value)

                                          
                     LABELS = ["Customer :","Number of Visitors :","Tag 1 :","Tag 2 :" ,"Tag 3 :"]

                     for x in range (len(LABELS)):
                             self.infodict[x].config(text = customerdetails[x])

                     tableSeatedCustomer[self.chosentablenum] = customerrow

                     print (tableSeatedCustomer)
                     

              def SeatVisitor(n):

                     CUSTOMERS = []

                     SelectCustomer = Toplevel(width=500, height =800 , bg='grey')

                     select = Label(SelectCustomer, text = "Please select a customer to seat. If the customer cannot be found please create one."
                                      ,pady = 10, width = 70,padx = 20,  bg = 'steelblue', fg = 'black' , relief = "raised" , borderwidth = 5)

                     select.pack()

                     self.chosentablenum = n
                     
                     import openpyxl

                     theFile = openpyxl.load_workbook('Customer Details Form.xlsx')
                     currentSheet = theFile['Sheet1']
                     first_column = currentSheet['A']

                     #Program iterates through names and adds them to a list to be accessed by the menu
                     for x in range(len(first_column)):
                            if currentSheet.cell(row=x+1, column=1).value == None:
                                   pass
                            else:
                                   CUSTOMERS.append(first_column[x].value) 

                     #Removes the header of the column (Names) from the choices
                     CUSTOMERS.remove("Name")
                     
                     variable = StringVar()
                     CustomerOptions = OptionMenu(SelectCustomer, variable, *CUSTOMERS, command = ChosenCustomer).pack()

                      
                     tabledict[self.chosentablenum].configure(bg = "red")
                     self.SeatVistiorBtn.config(text = "Unseat Visitor on Table", command=lambda n=n: UnseatVisitor(n))

                     

              #Function that will activate when a table has been clciked. It will open a toplevel window
              #and display info about the table
              def tableInfo (n):

                     self.tableInformation = Toplevel(width=500, height =800 , bg='grey')

                     tablenum = Label(self.tableInformation, text = "Table " + n, font = "Verdana 20", pady = 10, width = 20,
                                       padx = 20,  bg = 'steelblue'  , fg = 'black' , relief = "raised" , borderwidth = 5)

                     tablecapacitylbl = Label(self.tableInformation, text = "Seating Capacity :" , pady = 10, width = 20,
                                       padx = 20, bg = 'grey' , fg = 'black' , borderwidth = 5)

                     tablecapacity = Label(self.tableInformation, text = str(newvalue[n]) + " Visitors", pady = 10, width = 20,
                                       padx = 20, bg = 'grey' , fg = 'black' , borderwidth = 5)

                     tablestatuslbl = Label(self.tableInformation, text = "The table is :", pady = 10, width = 20,
                                       padx = 20, bg = 'grey' , fg = 'black' , borderwidth = 5)

                     tablestatus = Label(self.tableInformation, text = self.tablestatus[n], pady = 10, width = 20,
                                       padx = 20, bg = 'grey' , fg = 'black' , borderwidth = 5)

                     self.SeatVistiorBtn = Button(self.tableInformation, text = "Seat Visitor on Table", pady = 10, width = 20,padx = 20,
                                             bg = 'steelblue', fg = 'black' , borderwidth = 5, command=lambda n=n: SeatVisitor(n) )

                     CheckReservBtn = Button(self.tableInformation, text = "Check for Reservations", pady = 10, width = 20,padx = 20,
                                             bg = 'steelblue', fg = 'black' , borderwidth = 5, command=lambda n=n: CheckReservation(n) )


                     LABELS = ["Customer :","Number of Visitors :","Tag 1 :","Tag 2 :" ,"Tag 3 :"]

                     labeldict = {}
                     self.infodict = {}

                     for x in range (len(LABELS)):
                            labeldict[x] = Label(self.tableInformation, text = LABELS[x] , pady = 10, width = 20,
                                       padx = 20, bg = 'grey' , fg = 'black' , borderwidth = 5)
                            labeldict[x].grid(row = 4+x, column = 0)

                     for x in range (len(LABELS)):
                             self.infodict[x] = Label(self.tableInformation,pady = 10, width = 20,padx = 20,bg = 'grey',fg = 'black',borderwidth = 5)
                             self.infodict[x].grid(row = 4+x, column = 1)

                  
                     tablenum.grid(row=0,column=0,columnspan=2)
                     tablecapacitylbl.grid(row=1,column=0)
                     tablecapacity.grid(row=1,column=1)
                     tablestatuslbl.grid(row=2,column=0)
                     tablestatus.grid(row=2,column=1)
                     self.SeatVistiorBtn.grid(row=3,column=0)
                     CheckReservBtn.grid(row=3,column=1)

                     tablestatus = tabledict[n].cget('bg')

                     import openpyxl

                     theFile = openpyxl.load_workbook('Customer Details Form.xlsx')
                     currentSheet = theFile['Sheet1']


                     if tablestatus == "red":
                            self.SeatVistiorBtn.config(text = "Unseat Visitor on Table", command=lambda n=n: UnseatVisitor(n))

                            for x in tableSeatedCustomer:
                                    if x == str(n) :
                                          print (n)

                                          customerRow = tableSeatedCustomer[n]

                                          customerdetails = []

                                          for j in range(1, currentSheet.max_column+1):
                                                 customerdetails.append(currentSheet.cell(row=customerRow, column=j).value)

                                          for x in range (len(LABELS)):
                                                  self.infodict[x].config(text = customerdetails[x])

                     if tablestatus == "purple":

                            import openpyxl

                            theFile = openpyxl.load_workbook('ReservationForm.xlsx')
                            currentSheet = theFile['Sheet1']

                            RESERVATIONLBLS = ["Customer :","Number of Visitors :", "Time of Reservation:"]
                            
                            self.SeatVistiorBtn.config(text = "Seat Reserved Customer", command=lambda n=n: SeatReserved(n))

                            for x in range (len(LABELS)):
                                   labeldict[x].grid_forget()

                            for x in range (len(RESERVATIONLBLS)):
                                   labeldict[x] = Label(self.tableInformation, text = RESERVATIONLBLS[x] , pady = 10, width = 20,
                                       padx = 20, bg = 'grey' , fg = 'black' , borderwidth = 5)
                                   labeldict[x].grid(row = 4+x, column = 0)


                            reservCustomerDetails = []

                            customerRow = tableReservedCustomer[int(n)]

                            reservCustomerDetails.append(currentSheet.cell(row=customerRow, column=1).value)
                            reservCustomerDetails.append(currentSheet.cell(row=customerRow, column=2).value)
                            reservCustomerDetails.append(currentSheet.cell(row=customerRow, column=4).value)

                            for x in range (len(RESERVATIONLBLS)):
                                    self.infodict[x].config(text = reservCustomerDetails[x])
                            
                  

              import glob
              import os

              files = glob.glob('C:/Users/Akhil/Desktop/Python/Computing Project/**/*.txt',recursive = True)
              count = 0
              for x in files: 
                  count +=1 

              if count == 0:

                     def returntotablemenu():
                            self.master.show_frame(TableManager)
                            noPlans.destroy()
                            
                     noPlans = Toplevel(width=800, height =800 , bg='grey')
                     noFloors = Label(noPlans, text = "No saved floor plans. Create one, then return if needed.",
                                         bg = 'steelblue',width = 70, font=("Courier New", 10))
                     noFloors.grid(row = 0, column = 0 , columnspan = 4)

                     returntomenu = Button (noPlans , text = "Return to Menu" , command= returntotablemenu, bg = 'steelblue')
                     returntomenu.grid(row=2,column=0,columnspan=4)

              else:
                     from tkinter.filedialog import askopenfilename #Allows us to use tkinter file management
                     # show an "Open" dialog box and return the path to the selected file
                     self.filename = askopenfilename(initialdir='C:/Users/Akhil/Desktop/Python/Computing Project' , filetypes = [("Text Files", ".txt")]) 
                     
                     #Opens the chosen floor plan file and reads the colour of each table and saves it to a list
                     textfile = open(self.filename,"r") 
                     with open(self.filename) as f:
                         seatingcolourlist = f.read().splitlines()

                     seatingcolourdict = {} #Dictionary holding table and corresponding colour
                     tabledict = {} #Dictionary holding all the tables and their functions/information
                     
                     count1 = 1 # This iterates through the list and converts it to a dictionary
                     for x in seatingcolourlist :
                            seatingcolourdict[str(count1)] = x
                            count1+=1

                     info = Label (self, text = "Click a table to manage it ", font = "Verdana 15", padx = 10, pady = 10,
                                   bg = 'steelblue'  , fg = 'black' , relief = "raised" , borderwidth = 5)
                     info.grid(row=0,column=0, columnspan = 3) # Creates a label to inform the user

                     refreshbtn = Button(self, text = "Click to refresh tables ", font = "Verdana 12", padx = 10, pady = 10,
                                   bg = 'steelblue'  , fg = 'black' , relief = "raised" , borderwidth = 5, command = Refresh)
                     refreshbtn.grid(row=0,column=2, columnspan = 3)
                     

                     #Creates key of colours and corresponding seating capacity
                     key = Label (self, text = "Key: ", font = "Verdana 15" ,  bg = 'steelblue',
                                  fg = 'black' , relief = "raised" ).grid(row=1,column=0)

                     available= Label (self, text = "Available ", font = "Verdana 15" , bg = "green", padx = 5).grid(row=1,column=1,pady=(10,10))
                     reserved= Label (self, text = "Reserved ", font = "Verdana 15" , bg = "purple", padx = 5).grid(row=1,column=2,pady=(10,10))
                     occupied= Label (self, text = "Occupied ", font = "Verdana 15" , bg = "red", padx = 5).grid(row=1,column=3,pady=(10,10))
              
                     rowNumber = 3
                     columnNumber = 0

                     colourvalues = {"yellow" : "2" , "orange" : "4" , "red" : "6" , "cyan" : "10"}
                     status = {"green" : "Available", "purple" : "Reserved", "red" : "Occupied"}

                     #Converts colour into seating capacity 
                     newvalue = {}
                     count2 = 1
                     for x in seatingcolourlist :
                            z = colourvalues[x]  
                            newvalue[str(count2)] = z
                            count2+=1

                     
                     #Creates the grid of tables, coloured according to the saved floorplan
                     for n in seatingcolourdict:
                            tabledict[n] = Button(self,text = "Table " + str(n) + ": \n \n" +str(newvalue[n]) + " Seater", bg = "green",
                                                  font = "Verdana 12" ,width = 10, height = 5,command=lambda n=n: tableInfo(n))
                            tabledict[n].grid(row = rowNumber, column = columnNumber, padx = 5, pady = 10)

                            columnNumber += 1
                            if columnNumber == 5:
                                   columnNumber = 0
                                   rowNumber += 1


                     #Button at bottom of screen to open options to leave floor plan creation
                     returntofloormenu = Button(self, text = "Return to floor menu" , pady = 10 , padx = 10, font = "Verdana 12",
                                              command = MenuReturn ,  bg = 'steelblue'  , fg = 'black' , relief = "raised" , borderwidth = 5 )
                     returntofloormenu.grid(row = 10 , column = 1 , columnspan = 3)


                     Refresh()

                     #Converts colour into status
                     self.tablestatus = {}
                     count3 = 1
                     for x in tabledict :
                            y = tabledict[x].cget('bg')
                            z = status[y]
                            self.tablestatus[x] = z
                     

                     
                     #Small fix to keep GUI looking clean
                     self.grid_columnconfigure(1, weight=1)
                     self.grid_columnconfigure(3, weight=1)

                            


class LastTableManaging(Frame):
       def __init__(self, parent, controller):
              Frame.__init__(self, parent)
              Frame.config(self, bg= "grey")




              


#This handles the Automated Emails Functions               
class AutomatedEmails(Frame):
       def __init__(self, parent, controller):
              Frame.__init__(self, parent)
              Frame.config(self, bg= "grey")

              # this takes the information passed and forms the title label introducing the program
              def TitleText(char_ , x_ , y_ , span):

                     self.Greeting = Label(self, text = char_, width = 75, height = 5, bg = self.buttonbg , fg =self.buttonfg , borderwidth = 5, relief =self.buttonrelief)
                     self.Greeting.grid(row = x_, column = y_ , columnspan = span)

              
              # this takes the information passed and uses it to create the Menu Buttons
              def MenuButtons(char_, x_, y_ , n,z):
                     
                     self.b = Button(self, text = char_, width = self.button_width, height = 3, bg = self.buttonbg , fg =self.buttonfg , pady = 10 ,command=lambda: self.master.CreateFunctions(n,z))
                     self.b.grid(row=x_, column=y_ , pady = (5,5))


              # basic reuable button/label configuration
              self.button_width = 20
              self.buttonbg = 'steelblue'
              self.buttonfg = 'black'
              self.buttonrelief= 'raised'
              
              # the information to be passed into the label - text , row , column , columnspan
              TitleText("Automated Emails Management Menu \n \n Choose what you want to do!", 0 ,0,3)


              #the information to be passed into the buttons - text , row , column , columnspan and the next screen to navigate to
              MenuButtons('Create New Email to Send', 2, 1 , CreateEmail,CreateEmail)
              MenuButtons('View and Edit\n Saved Emails', 3, 1, EmailEditor,EmailEditor)
              MenuButtons('Return to Menu', 4, 1, MainMenu,MainMenu)

              self.grid_columnconfigure(1, weight=1)


class CreateEmail(Frame):
       def __init__(self, parent, controller):
              Frame.__init__(self, parent)
              Frame.config(self, bg= "grey")
              self.grid_columnconfigure(1, weight=1)

              def sendEmail():
                     self.confirmation.destroy ()
                     OptInorCustom = self.r.get() #Whether the user is sending to all opted in or custom choices
                     text=self.box.get("1.0","end") #Reads the email created in the text box

                     if OptInorCustom == 1: #If email being sent to all customers that opted in

                            import openpyxl #Excel file manager module

                            SendTo = []#A list of everyone to send the email to

                            #Opens excel file and sheet containg database and selects the fist column (customer names)
                            theFile = openpyxl.load_workbook('Customer Details Form.xlsx')
                            currentSheet = theFile['Sheet1']
                            optcolumn = currentSheet['D']

                            #Program iterates through names and adds them to a list to be accessed by the menu
                            for x in range(0,len(optcolumn)):
                                   if currentSheet.cell(row=x+1, column=4).value == 1:
                                          z = currentSheet.cell(row=x+1, column=3).value
                                          SendTo.append(z)  #Add to list of recipients                                        

                     if OptInorCustom == 2:#If email being sent to a custom choice of customeres

                            #Opens the customer details form excel sheet
                            import openpyxl
                            theFile = openpyxl.load_workbook('Customer Details Form.xlsx')
                            currentSheet = theFile['Sheet1']
                            names = currentSheet['A']

                            SendTo = [] #A list of everyone to send the email to
                            
                            for x in range (0, self.NumOfRecipients): #Iterate through the amount of entry boxes created
                                   z = self.RecepientsDict[x].get()

                                   if "@" in z:                #If an email is entered just use it
                                          SendTo.append(z)
                                          
                                   else:                       #If a name is entered check if it is in the customer database
                                          for y in range(0,len(names)):
                                                 if currentSheet.cell(row=y+1, column=1).value == z:
                                                        emailFromName = currentSheet.cell(row=y+1, column=3).value
                                                        SendTo.append(emailFromName)#Add to list of recipients
                                                 
                                   InvalidEmail()

                     import smtplib # Imports the email handling module

                     #Email address and password of the program
                     EMAIL_ADDRESS = 'rmanagertest@gmail.com' 
                     EMAIL_PASSWORD = 'Testing123%'

                     with smtplib.SMTP('smtp.gmail.com' , 587) as smtp: #Sets it as a gmail account

                            #Sets up encryption
                            smtp.ehlo()
                            smtp.starttls()
                            smtp.ehlo()

                            smtp.login(EMAIL_ADDRESS , EMAIL_PASSWORD) # Logs into account using details

                            #Creates the email
                            subject = "This is a Test"
                            body = text

                            msg = f'Subject: {subject} \n\n {body}' # Creates mail


                            for x in SendTo: 
                                   smtp.sendmail(EMAIL_ADDRESS , x , msg) #Sends the email

              #Error warning if the full name used to find an email doesn't exist                    
              def InvalidEmail():
                     nonEmail = Toplevel(width=500, height =800 , bg='grey')#Opens a new window to house the GUI
                     #Label to ask the user whether they are sure they want to reserve this seat
                     notFound = Label(nonEmail, text = "You have entered a name not in the customer database. Enter a correct one or an email address",pady = 10,
                                                padx = 20,  bg = 'steelblue'  , fg = 'black' , relief = "raised" , borderwidth = 5)
                     notFound.pack()               

              def sendConfirm():
                     self.confirmation = Toplevel(width=500, height =800 , bg='grey')#Opens a new window to house the GUI
                     #Label to ask the user whether they are sure they want to reserve this seat
                     sendmailLbl = Label(self.confirmation, text = "Are you sure you want to send this email to all the chosen parties?.",pady = 10,
                                                padx = 20,  bg = 'steelblue'  , fg = 'black' , relief = "raised" , borderwidth = 5)
                     sendmailLbl.pack()

                     confirmYes = Button(self.confirmation, text = "Yes - Send Mail",bg = 'steelblue', fg = 'black', command = sendEmail).pack()
                     confirmNo = Button(self.confirmation, text = "No - Don't Send Mail",bg = 'steelblue', fg = 'black', command = self.confirmation.destroy).pack()

              def ErrorInvalidNumber():
                     InvalidNum = Toplevel(width=500, height =800 , bg='grey')#Opens a new window to house the GUI
                     #Label to ask the user whether they are sure they want to reserve this seat
                     chooselabel = Label(InvalidNum,text = "You have entered an invalid number. Please enter one between 1 and 6.",pady = 10,
                                                padx = 20,  bg = 'steelblue'  , fg = 'black' , relief = "raised" , borderwidth = 5)
                     chooselabel.pack()

              def Recipients(): #Creates the designated amount of entry boxes to let the user enter full names of customers or their emails
                     self.count +=1
                     
                     self.NumOfRecipients = int(self.NumToSend.get()) # The entered number of entry boxes to create

                     if self.NumOfRecipients in range (1,7): # Checks if the numbr of boxes to create is valid

                            for x in range (0, self.NumOfRecipients): # Iterates through, creatiing the entry boxes and labels

                                   self.RecepientsLblDict[x] = Label(self, text= "Recipient " + str(x+1) + " :" , bg = 'grey', pady = 10 )
                                   self.RecepientsLblDict[x].grid(row = 6+x, column = 0, sticky = W)

                                   self.RecepientsDict[x] = Entry(self , width = 42, fg = 'white' , bg = 'black' , borderwidth = 10 )
                                   self.RecepientsDict[x].grid(row= 6+x, column = 1, columnspan = 2, pady = (10,10))
                                   
                     else:
                            ErrorInvalidNumber() #Creates error mssage if number is invalid
                            

              def AllOpted(value):

                     if self.count == 1: #Used to remove the entry box for how many recipients
                            
                            self.NumToSend.grid_forget()
                            self.ConfirmNum.grid_forget()
                            self.HowMany.grid_forget()

                     if self.count > 1:#Used to remove the entry box for how many recipients and any entry boxes created

                            self.NumToSend.grid_forget()
                            self.ConfirmNum.grid_forget()
                            self.HowMany.grid_forget()
                            
                            for x in range (0, len(self.RecepientsLblDict)):
                                   
                                   self.RecepientsLblDict[x].destroy()
                                   self.RecepientsDict[x].destroy()

                     self.count = 0
                     
              def Custom(value): # Creates an entry box to let the user choose how many recipients they want to send mails to
                     self.count +=1
                     self.HowMany = Label(self, text="How many customers are you \n sending an email to? (Max = 6)" , bg = 'grey', pady = 10 )
                     self.HowMany.grid(row = 4 , column = 0, sticky = W)

                     self.NumToSend = Entry(self , width = 42, fg = 'white' , bg = 'black' , borderwidth = 10 )
                     self.NumToSend.grid(row= 4, column = 1, columnspan = 2)
                     self.NumToSend.insert(0, "Enter number of recipients")

                     self.ConfirmNum = Button(self, text = "Continue", command= Recipients ,  bg = 'steelblue', pady = 10)
                     self.ConfirmNum.grid(row=5, column = 1, columnspan = 2)

              def SaveEmail():
                     import os.path
                   
                     emailtoSave=self.box.get("1.0","end")
                     name = self.SaveEntry.get()
                     
                     save_path = 'C:/Users/Akhil/Desktop/Python/Computing Project/SavedEmails/'

                     completeName = os.path.join(save_path, name+".txt")                     

                     with open(completeName, "w") as text_file:
                            text_file.write(emailtoSave)

                     self.master.CreateFunctions(CreateEmail,AutomatedEmails)
                     self.SaveName.destroy()

              def EmailName():
                     self.SaveorExit.destroy()

                     self.SaveName = Toplevel(width=500, height =800 , bg='grey')#Opens a new window to house the GUI
                     #Label to ask the user whether they are sure they want to reserve this seat
                     SaveNameLbl = Label(self.SaveName,text = "Enter the name you wish to save the email as",pady = 10,
                                                padx = 20,  bg = 'steelblue'  , fg = 'black' , relief = "raised" , borderwidth = 5)
                     SaveNameLbl.pack()

                     self.SaveEntry = Entry(self.SaveName, width = 42, fg = 'white' , bg = 'black' , borderwidth = 10)
                     self.SaveEntry.pack()

                     YesBtn = Button(self.SaveName,text = "Save Email",pady = 10,padx = 20,bg = 'steelblue',
                                     fg = 'black' , relief = "raised" , borderwidth = 5, command = SaveEmail)
                     YesBtn.pack()

              def ExitToMenu():
                     self.master.CreateFunctions(CreateEmail,AutomatedEmails)
                     self.SaveorExit.destroy()


              def SaveandExit():
                     self.SaveorExit = Toplevel(width=500, height =800 , bg='grey')#Opens a new window to house the GUI
                     #Label to ask the user whether they are sure they want to reserve this seat
                     Savequestion = Label(self.SaveorExit,text = "Do you wish to save the mail you have written. Yes or No?",pady = 10,
                                                padx = 20,  bg = 'steelblue'  , fg = 'black' , relief = "raised" , borderwidth = 5)
                     Savequestion.pack()

                     YesBtn = Button(self.SaveorExit,text = "Yes - Save email and Exit",pady = 10,padx = 20,bg = 'steelblue',
                                     fg = 'black' , relief = "raised" , borderwidth = 5, command = EmailName)
                     YesBtn.pack()

                     NoBtn = Button(self.SaveorExit,text = "No - Exit to Menu",pady = 10,padx = 20,bg = 'steelblue', fg = 'black' ,
                                    relief = "raised" , borderwidth = 5, command = ExitToMenu)
                     NoBtn.pack()
                                          

              #Dictionaries to hold labels and entry fields
              self.RecepientsLblDict = {}
              self.RecepientsDict = {}
              self.count = 0
                            
              TitleText = Label(self, text="Please fill out the form to sent the email",  bg = 'steelblue'  , fg = 'black' ,
                                borderwidth = 5, pady = 10 , padx = 10, relief= 'raised', font=("Courier New", 10))
              TitleText.grid(row = 0 , column = 0 , columnspan = 3)#Label at top of screen to intrtoduce function


              email = Label(self, text="Enter the email you want to send :" , bg = 'grey', pady = 30 , padx = 10)
              email.grid(row = 1 , column = 0, sticky = W)#Label informing the user where to write the email

              self.box = Text(self, height = 10 , width = 50, borderwidth = 2) #Creating the textbox to write the email in
              self.box.grid(row = 1, column = 1, columnspan = 2, pady = (10,10))
              self.box.insert(INSERT, "Enter email text here")

              #Creating the radiobuttons for whether the user wants to send to all those opted in or a custom choice of customers
              self.r = IntVar()
              self.r.set(1)
              OptIn = Label(self, text="Who to send the email to :" , bg = 'grey', pady = 30 , padx = 10).grid(row = 2 , column = 0, sticky = W)
              Radiobutton(self, text = "All Opted In" , variable = self.r , value = 1 , bg = 'grey', command = lambda: AllOpted(self.r.get())).grid(row= 2, column = 1, sticky = W)
              Radiobutton(self, text = "Custom" , variable = self.r , value = 2 ,  bg = 'grey', command = lambda: Custom(self.r.get())).grid(row= 2, column = 2, sticky = W , padx = (2,100))    

              self.submit1 = Button(self, text = "Send Email",command = sendConfirm,   bg = 'steelblue', pady = 10  ) #Button to send the email
              self.submit1.grid(row=15, column = 1, sticky=W )

              SaveandExit = Button(self, text = "Save and Exit to Menu",  bg = 'steelblue', pady = 8 , borderwidth = 4 ,command = SaveandExit )#Exit Button
              SaveandExit.grid(row=15, column = 2, sticky=W)


class EmailEditor(Frame):
       def __init__(self, parent, controller):
              Frame.__init__(self, parent)
              Frame.config(self, bg= "grey")
              self.grid_columnconfigure(1, weight=1)


              def sendEmail():
                     self.confirmation.destroy ()
                     OptInorCustom = self.r.get() #Whether the user is sending to all opted in or custom choices
                     text=self.box.get("1.0","end") #Reads the email created in the text box

                     if OptInorCustom == 1: #If email being sent to all customers that opted in

                            import openpyxl #Excel file manager module

                            SendTo = []#A list of everyone to send the email to

                            #Opens excel file and sheet containg database and selects the fist column (customer names)
                            theFile = openpyxl.load_workbook('Customer Details Form.xlsx')
                            currentSheet = theFile['Sheet1']
                            optcolumn = currentSheet['D']

                            #Program iterates through names and adds them to a list to be accessed by the menu
                            for x in range(0,len(optcolumn)):
                                   if currentSheet.cell(row=x+1, column=4).value == 1:
                                          z = currentSheet.cell(row=x+1, column=3).value
                                          SendTo.append(z)  #Add to list of recipients                                        

                     if OptInorCustom == 2:#If email being sent to a custom choice of customeres

                            #Opens the customer details form excel sheet
                            import openpyxl
                            theFile = openpyxl.load_workbook('Customer Details Form.xlsx')
                            currentSheet = theFile['Sheet1']
                            names = currentSheet['A']

                            SendTo = [] #A list of everyone to send the email to
                            
                            for x in range (0, self.NumOfRecipients): #Iterate through the amount of entry boxes created
                                   z = self.RecepientsDict[x].get()

                                   if "@" in z:                #If an email is entered just use it
                                          SendTo.append(z)
                                          
                                   else:                       #If a name is entered check if it is in the customer database
                                          for y in range(0,len(names)):
                                                 if currentSheet.cell(row=y+1, column=1).value == z:
                                                        emailFromName = currentSheet.cell(row=y+1, column=3).value
                                                        SendTo.append(emailFromName)#Add to list of recipients
                                                 
                                   InvalidEmail()

                     import smtplib # Imports the email handling module

                     #Email address and password of the program
                     EMAIL_ADDRESS = 'rmanagertest@gmail.com' 
                     EMAIL_PASSWORD = 'Testing123%'

                     with smtplib.SMTP('smtp.gmail.com' , 587) as smtp: #Sets it as a gmail account

                            #Sets up encryption
                            smtp.ehlo()
                            smtp.starttls()
                            smtp.ehlo()

                            smtp.login(EMAIL_ADDRESS , EMAIL_PASSWORD) # Logs into account using details

                            #Creates the email
                            subject = "This is a Test"
                            body = text

                            msg = f'Subject: {subject} \n\n {body}' # Creates mail


                            for x in SendTo: 
                                   smtp.sendmail(EMAIL_ADDRESS , x , msg) #Sends the email

              #Error warning if the full name used to find an email doesn't exist                    
              def InvalidEmail():
                     nonEmail = Toplevel(width=500, height =800 , bg='grey')#Opens a new window to house the GUI
                     #Label to ask the user whether they are sure they want to reserve this seat
                     notFound = Label(nonEmail, text = "You have entered a name not in the customer database. Enter a correct one or an email address",pady = 10,
                                                padx = 20,  bg = 'steelblue'  , fg = 'black' , relief = "raised" , borderwidth = 5)
                     notFound.pack()               

              def sendConfirm():
                     self.confirmation = Toplevel(width=500, height =800 , bg='grey')#Opens a new window to house the GUI
                     #Label to ask the user whether they are sure they want to reserve this seat
                     sendmailLbl = Label(self.confirmation, text = "Are you sure you want to send this email to all the chosen parties?.",pady = 10,
                                                padx = 20,  bg = 'steelblue'  , fg = 'black' , relief = "raised" , borderwidth = 5)
                     sendmailLbl.pack()

                     confirmYes = Button(self.confirmation, text = "Yes - Send Mail",bg = 'steelblue', fg = 'black', command = sendEmail).pack()
                     confirmNo = Button(self.confirmation, text = "No - Don't Send Mail",bg = 'steelblue', fg = 'black', command = self.confirmation.destroy).pack()

              def ErrorInvalidNumber():
                     InvalidNum = Toplevel(width=500, height =800 , bg='grey')#Opens a new window to house the GUI
                     #Label to ask the user whether they are sure they want to reserve this seat
                     chooselabel = Label(InvalidNum,text = "You have entered an invalid number. Please enter one between 1 and 6.",pady = 10,
                                                padx = 20,  bg = 'steelblue'  , fg = 'black' , relief = "raised" , borderwidth = 5)
                     chooselabel.pack()

              def Recipients(): #Creates the designated amount of entry boxes to let the user enter full names of customers or their emails
                     self.count +=1
                     
                     self.NumOfRecipients = int(self.NumToSend.get()) # The entered number of entry boxes to create

                     if self.NumOfRecipients in range (1,7): # Checks if the numbr of boxes to create is valid

                            for x in range (0, self.NumOfRecipients): # Iterates through, creatiing the entry boxes and labels

                                   self.RecepientsLblDict[x] = Label(self, text= "Recipient " + str(x+1) + " :" , bg = 'grey', pady = 10 )
                                   self.RecepientsLblDict[x].grid(row = 6+x, column = 0, sticky = W)

                                   self.RecepientsDict[x] = Entry(self , width = 42, fg = 'white' , bg = 'black' , borderwidth = 10 )
                                   self.RecepientsDict[x].grid(row= 6+x, column = 1, columnspan = 2, pady = (10,10))
                                   
                     else:
                            ErrorInvalidNumber() #Creates error mssage if number is invalid                            

              def AllOpted(value):

                     if self.count == 1: #Used to remove the entry box for how many recipients
                            
                            self.NumToSend.grid_forget()
                            self.ConfirmNum.grid_forget()
                            self.HowMany.grid_forget()

                     if self.count > 1:#Used to remove the entry box for how many recipients and any entry boxes created

                            self.NumToSend.grid_forget()
                            self.ConfirmNum.grid_forget()
                            self.HowMany.grid_forget()
                            
                            for x in range (0, len(self.RecepientsLblDict)):
                                   
                                   self.RecepientsLblDict[x].destroy()
                                   self.RecepientsDict[x].destroy()

                     self.count = 0
                     
              def Custom(value): # Creates an entry box to let the user choose how many recipients they want to send mails to
                     self.count +=1
                     self.HowMany = Label(self, text="How many customers are you \n sending an email to? (Max = 6)" , bg = 'grey', pady = 10 )
                     self.HowMany.grid(row = 4 , column = 0, sticky = W)

                     self.NumToSend = Entry(self , width = 42, fg = 'white' , bg = 'black' , borderwidth = 10 )
                     self.NumToSend.grid(row= 4, column = 1, columnspan = 2)
                     self.NumToSend.insert(0, "Enter number of recipients")

                     self.ConfirmNum = Button(self, text = "Continue", command= Recipients ,  bg = 'steelblue', pady = 10)
                     self.ConfirmNum.grid(row=5, column = 1, columnspan = 2)

              def SaveEmail():

                     import os.path
                   
                     emailtoSave=self.box.get("1.0","end")
                     name = self.SaveEntry.get()
                     
                     save_path = 'C:/Users/Akhil/Desktop/Python/Computing Project/SavedEmails/'

                     completeName = os.path.join(save_path, name)                     

                     with open(completeName, "w") as text_file:
                            text_file.write(emailtoSave)

                     self.master.CreateFunctions(CreateEmail,AutomatedEmails)
                     self.SaveName.destroy()

              def EmailName():
                     self.SaveorExit.destroy()

                     self.SaveName = Toplevel(width=500, height =800 , bg='grey')#Opens a new window to house the GUI
                     #Label to ask the user whether they are sure they want to reserve this seat
                     SaveNameLbl = Label(self.SaveName,text = "Change name to save as new or click save to overwrite",pady = 10,
                                                padx = 20,  bg = 'steelblue'  , fg = 'black' , relief = "raised" , borderwidth = 5)
                     SaveNameLbl.pack()

                     self.SaveEntry = Entry(self.SaveName, width = 42, fg = 'white' , bg = 'black' , borderwidth = 10)
                     self.SaveEntry.insert(0, os.path.basename(self.filename))
                     self.SaveEntry.pack()

                     YesBtn = Button(self.SaveName,text = "Save Email",pady = 10,padx = 20,bg = 'steelblue',
                                     fg = 'black' , relief = "raised" , borderwidth = 5, command = SaveEmail)
                     YesBtn.pack()

              def ExitToMenu():
                     self.master.CreateFunctions(CreateEmail,AutomatedEmails)
                     self.SaveorExit.destroy()

              def SaveandExit():
                     self.SaveorExit = Toplevel(width=500, height =800 , bg='grey')#Opens a new window to house the GUI
                     #Label to ask the user whether they are sure they want to reserve this seat
                     Savequestion = Label(self.SaveorExit,text = "Do you wish to save the mail you have written. Yes or No?",pady = 10,
                                                padx = 20,  bg = 'steelblue'  , fg = 'black' , relief = "raised" , borderwidth = 5)
                     Savequestion.pack()

                     YesBtn = Button(self.SaveorExit,text = "Yes - Save email and Exit",pady = 10,padx = 20,bg = 'steelblue',
                                     fg = 'black' , relief = "raised" , borderwidth = 5, command = EmailName)
                     YesBtn.pack()

                     NoBtn = Button(self.SaveorExit,text = "No - Exit to Menu",pady = 10,padx = 20,bg = 'steelblue', fg = 'black' ,
                                    relief = "raised" , borderwidth = 5, command = ExitToMenu)
                     NoBtn.pack()

              import glob
              import os

              files = glob.glob('C:/Users/Akhil/Desktop/Python/Computing Project/SavedEmails/**/*.txt',recursive = True)
              count = 0
              for x in files: 
                  count +=1 

              if count == 0:

                     def returntoEmailmenu():
                            self.master.CreateFunctions(CreateEmail,AutomatedEmails)
                            noEmailsSaved.destroy()
                            

                     noEmailsSaved = Toplevel(width=800, height =800 , bg='grey')
                     noEmails = Label(noEmailsSaved, text = "No saved floor plans. Create one, then return if needed.",
                                         bg = 'steelblue',width = 70, font=("Courier New", 10))
                     noEmails.grid(row = 0, column = 0 , columnspan = 4)

                     returntomenu = Button (noEmailsSaved , text = "Return to Menu" , command= returntoEmailmenu, bg = 'steelblue')
                     returntomenu.grid(row=2,column=0,columnspan=4)

              else:
 
                     from tkinter.filedialog import askopenfilename #Allows us to use tkinter file management
                     # show an "Open" dialog box and return the path to the selected file
                     self.filename = askopenfilename(initialdir='C:/Users/Akhil/Desktop/Python/Computing Project/SavedEmails/' , filetypes = [("Text Files", ".txt")]) 
                     
                     #Opens the chosen floor plan file and reads the colour of each table and saves it to a list
                     textfile = open(self.filename,"r") 
                     with open(self.filename) as f:
                         text = f.read()

              #Dictionaries to hold labels and entry fields
              self.RecepientsLblDict = {}
              self.RecepientsDict = {}
              self.count = 0
                            
              TitleText = Label(self, text="This is the email you selected - Edit it as you wish",  bg = 'steelblue'  , fg = 'black' ,
                                borderwidth = 5, pady = 10 , padx = 10, relief= 'raised', font=("Courier New", 10))
              TitleText.grid(row = 0 , column = 0 , columnspan = 3)#Label at top of screen to intrtoduce function


              email = Label(self, text="Edit email :" , bg = 'grey', pady = 30 , padx = 10)
              email.grid(row = 1 , column = 0, sticky = W)#Label informing the user where to write the email

              self.box = Text(self, height = 10 , width = 50, borderwidth = 2) #Creating the textbox to write the email in
              self.box.grid(row = 1, column = 1, columnspan = 2, pady = (10,10))
              self.box.insert(INSERT, text)

              #Creating the radiobuttons for whether the user wants to send to all those opted in or a custom choice of customers
              self.r = IntVar()
              self.r.set(1)
              OptIn = Label(self, text="Who to send the email to :" , bg = 'grey', pady = 30 , padx = 10).grid(row = 2 , column = 0, sticky = W)
              Radiobutton(self, text = "All Opted In" , variable = self.r , value = 1 , bg = 'grey', command = lambda: AllOpted(self.r.get())).grid(row= 2, column = 1, sticky = W)
              Radiobutton(self, text = "Custom" , variable = self.r , value = 2 ,  bg = 'grey', command = lambda: Custom(self.r.get())).grid(row= 2, column = 2, sticky = W , padx = (2,100))    

              self.submit1 = Button(self, text = "Send Email",command = sendConfirm,   bg = 'steelblue', pady = 10  ) #Button to send the email
              self.submit1.grid(row=15, column = 1, sticky=W )

              SaveandExit = Button(self, text = "Save As and Exit to Menu",  bg = 'steelblue', pady = 8 , borderwidth = 4 ,command = SaveandExit )#Exit Button
              SaveandExit.grid(row=15, column = 2, sticky=W)


#This handles the Floor Plan Creator Function
class FloorPlanCreator(Frame):
       def __init__(self, parent, controller):
              Frame.__init__(self, parent)
              Frame.config(self, bg= "grey")

              #Function that will activate when a table has been clciked. It will open a toplevel window
              #and display buttons allowing the user to choose seating capacity
              def seatchoice (n):

                  seatchoice = Toplevel(width=500, height =800 , bg='grey')

                  chooselabel = Label (seatchoice, text = "Choose the seating capacity for : " + n, font = "Verdana 20", pady = 10,
                                       padx = 20,  bg = 'steelblue'  , fg = 'black' , relief = "raised" , borderwidth = 5)
                  chooselabel.grid(row=0,column=0,columnspan=3)

                  twoseaterbtn = Button (seatchoice, text = "2 Seats ", font = "Verdana 15" , bg = "yellow", pady = 10,command=lambda n=n:selected(n , "yellow"))
                  twoseaterbtn.grid(row=1,column=1)

                  fourseaterbtn = Button (seatchoice, text = "4 Seats ", font = "Verdana 15" , bg = "orange", pady = 10,command=lambda n=n:selected(n , "orange"))
                  fourseaterbtn.grid(row=2,column=1)

                  sixseaterbtn = Button (seatchoice, text = "6 Seats ", font = "Verdana 15" , bg = "red", pady = 10,command=lambda n=n:selected(n , "red"))
                  sixseaterbtn.grid(row=3,column=1)

                  tenseaterbtn = Button (seatchoice, text = "10 Seats ", font = "Verdana 15" , bg = "cyan", pady = 10,command=lambda n=n:selected(n, "cyan"))
                  tenseaterbtn.grid(row=4,column=1)

              #Function that activates once the user has selected a table capacity and print it to console
              #as well as places it in the correct space in seatingtabledict and changes th ecolour in the creator
              def selected (n , x):
                  seatingcolourdict[n] = x
                  tabledict[n].configure(bg = x)
                  
              def leavecreationwarning():
                     #Creates a top level window to hold the exit box
                     Warning = Toplevel(bg = "gray")

                     #A label warning the the user data may be lost
                     msg = Label(Warning, text="Are you sure you want to return to Floor Plan Menu? Current entries may be deleted or saved for later.",
                                 bg = 'steelblue'  , fg = 'black', relief= 'raised', borderwidth = 5, pady = 15 , padx = 10,)
                     msg.pack()

                     #Closes the exit box
                     def CreationReturn():

                            Warning.destroy()

                     #Uses the master function to reset the floor plan creation function and return to floor plan menu
                     def FloorMenuReturn():

                            self.master.CreateFunctions(FloorPlanMenu,FloorPlanMenu)
                            Warning.destroy()
                     
                     #Returns user to floor plan menu but keeps entered data in box to be returned to
                     def TempFloorMenuReturn():

                            self.master.show_frame(FloorPlanMenu)
                            Warning.destroy()

                     #Creates the button to close the exit box
                     CreationReturnBtn = Button(Warning, text="Return to Floor Plan Creation", command= CreationReturn,
                                                bg = 'steelblue'  , fg = 'black', relief= 'raised', borderwidth = 5, pady = 5 , padx = 10,)
                     CreationReturnBtn.pack()

                     #Creates the button to return to menu and clear entries
                     FloorMenuReturnBtn = Button(Warning, text="Return to Floor Plan Menu and Clear Changes", command= FloorMenuReturn,
                                                 bg = 'steelblue'  , fg = 'black', relief= 'raised', borderwidth = 5, pady = 5 , padx = 10,)
                     FloorMenuReturnBtn.pack()

                     #Creates the button to return to menu but keep entries
                     TempFloorMenuBtn = Button(Warning, text="Return to Menu But Keep Changes Saved",command= TempFloorMenuReturn,
                                               bg = 'steelblue'  , fg = 'black', relief= 'raised', borderwidth = 5, pady = 5 , padx = 10,)
                     TempFloorMenuBtn.pack()
                  

              #Function that activates when save to file button pressed - uses entered file name to to create/open file
              #and enters the table colours into it to be used by table manager
              def saveandexit():

                  filename = self.entrybox.get()

                  textfile = open(filename + ".txt","w") 
                  for z in seatingcolourdict :
                      textfile.write(seatingcolourdict[z] + "\n")
                           

              #Function that is activated when save and exit button pressed - it opens a new toplevel window containing
              #an entry box to enter file name to be saved as as well as a button to activate saveandexit()
              def nameenter():

                  saving = Toplevel(width=500, height =800 , bg='grey')
                  self.entrybox = Entry (saving , width = 52, fg = 'white' , bg = 'black' , borderwidth = 10)
                  self.entrybox.grid(row=0,column=0,columnspan=3)
                  self.entrybox.insert(0 , "Enter name of floor plan to be saved")

                  submitbtn = Button (saving , text = "Save to File" , command= saveandexit , bg = 'steelblue'). grid(row=1,column=0,columnspan=3)
              

              tables = []

              for x in range (1, 26):
                  tablename = ("Table " + str(x))
                  tables.append(tablename)

              tabledict = {}
              seatingcolourdict = {}

              rowNumber = 2
              columnNumber = 0

              info = Label (self, text = "Click a table to set its seating capacity to match your restaurant ",font = "Verdana 15",
                            padx = 10, pady = 10,  bg = 'steelblue'  , fg = 'black' , relief = "raised" , borderwidth = 5)
              info.grid(row=0,column=0, columnspan = 5, sticky = E)

              key = Label (self, text = "Key: ", font = "Verdana 15" ,  bg = 'steelblue'  , fg = 'black' , relief = "raised" ).grid(row=1,column=0)
              twoseater = Label (self, text = "2 Seats ", font = "Verdana 15" , bg = "yellow", padx = 5).grid(row=1,column=1)
              fourseater = Label (self, text = "4 Seats ", font = "Verdana 15" , bg = "orange", padx = 5).grid(row=1,column=2)
              sixseater = Label (self, text = "6 Seats ", font = "Verdana 15" , bg = "red", padx = 5).grid(row=1,column=3)
              tenseater = Label (self, text = "10 Seats ", font = "Verdana 15" , bg = "cyan", padx = 5).grid(row=1,column=4)

              #Iterates through number of tables and creates a command to use them and stores it in the tabledict dictionary
              #as well as setting the colour value for each table as white ("Not Set")
              for n in tables:

                  tabledict[n] = Button(self,text=n,width = 10, height = 5,command=lambda n=n: seatchoice(n))
                  tabledict[n].grid(row = rowNumber, column = columnNumber, padx = 5, pady = 10)

                  seatingcolourdict[n] = 'Not Set'
                  
                  columnNumber += 1
                  if columnNumber == 5:
                      columnNumber = 0
                      rowNumber += 1

              #Button at bottom of screen to open options to leave floor plan creation
              returntofloormenu = Button(self, text = "Return to floor menu" , pady = 10 , padx = 10, font = "Verdana 12",
                                       command = leavecreationwarning ,  bg = 'steelblue'  , fg = 'black' , relief = "raised" , borderwidth = 5 )
              returntofloormenu.grid(row = 7 , column = 0 , columnspan = 2)

              #Button at bottom of screen to run nameenter function and then save and exit
              saveandexitbtn = Button (self, text = "Save Floor Plan and Exit to Menu" , pady = 10 , padx = 10, font = "Verdana 12",
                                       command =nameenter ,  bg = 'steelblue'  , fg = 'black' , relief = "raised" , borderwidth = 5 )
              saveandexitbtn.grid(row = 7 , column = 2 , columnspan = 3)


#This handles the Floor Plan Creator Function
class FloorPlanEditor(Frame):
       def __init__(self, parent, controller):
              Frame.__init__(self, parent)
              Frame.config(self, bg= "grey")


              def seatchoice (n):

                  seatchoice = Toplevel(width=500, height =800 , bg='grey')

                  chooselabel = Label (seatchoice, text = "Choose the seating capacity for : " + n, font = "Verdana 20", pady = 10,
                                       padx = 20,  bg = 'steelblue'  , fg = 'black' , relief = "raised" , borderwidth = 5)
                  chooselabel.grid(row=0,column=0,columnspan=3)

                  twoseaterbtn = Button (seatchoice, text = "2 Seats ", font = "Verdana 15" , bg = "yellow", pady = 10,command=lambda n=n:selected(n , "yellow"))
                  twoseaterbtn.grid(row=1,column=1)

                  fourseaterbtn = Button (seatchoice, text = "4 Seats ", font = "Verdana 15" , bg = "orange", pady = 10,command=lambda n=n:selected(n , "orange"))
                  fourseaterbtn.grid(row=2,column=1)

                  sixseaterbtn = Button (seatchoice, text = "6 Seats ", font = "Verdana 15" , bg = "red", pady = 10,command=lambda n=n:selected(n , "red"))
                  sixseaterbtn.grid(row=3,column=1)

                  tenseaterbtn = Button (seatchoice, text = "10 Seats ", font = "Verdana 15" , bg = "cyan", pady = 10,command=lambda n=n:selected(n, "cyan"))
                  tenseaterbtn.grid(row=4,column=1)

              #Function that activates once the user has selected a table capacity and print it to console
              #as well as places it in the correct space in seatingtabledict and changes th ecolour in the creator
              def selected (n , x):
                  seatingcolourdict[n] = x
                  tabledict[n].configure(bg = x)
                  
              def leavecreationwarning():
                     #Creates a top level window to hold the exit box
                     Warning = Toplevel(bg = "gray")

                     #A label warning the the user data may be lost
                     msg = Label(Warning, text="Are you sure you want to return to Floor Plan Menu? Unsaved changes will be deleted.",
                                 bg = 'steelblue'  , fg = 'black', relief= 'raised', borderwidth = 5, pady = 15 , padx = 10,)
                     msg.pack()

                     #Closes the exit box
                     def CreationReturn():

                            Warning.destroy()

                     #Uses the master function to reset the floor plan creation function and return to floor plan menu
                     def FloorMenuReturn():

                            self.master.CreateFunctions(FloorPlanMenu,FloorPlanMenu)
                            Warning.destroy()

                     #Creates the button to close the exit box
                     CreationReturnBtn = Button(Warning, text="Return to Floor Plan Creation", command= CreationReturn,
                                                bg = 'steelblue'  , fg = 'black', relief= 'raised', borderwidth = 5, pady = 5 , padx = 10,)
                     CreationReturnBtn.pack()

                     #Creates the button to return to menu and clear entries
                     FloorMenuReturnBtn = Button(Warning, text="Return to Floor Plan Menu and Clear Changes", command= FloorMenuReturn,
                                                 bg = 'steelblue'  , fg = 'black', relief= 'raised', borderwidth = 5, pady = 5 , padx = 10,)
                     FloorMenuReturnBtn.pack()
                  

              #Function that activates when save to file button pressed - uses entered file name to to create/open file
              #and enters the table colours into it to be used by table manager
              def saveandexit():

                  filename = self.entrybox.get()

                  textfile = open(filename,"w") 
                  for z in seatingcolourdict :
                      textfile.write(seatingcolourdict[z] + "\n")
                           

              #Function that is activated when save and exit button pressed - it opens a new toplevel window containing
              #an entry box to enter file name to be saved as as well as a button to activate saveandexit()
              def nameenter():
                  import os

                  saving = Toplevel(width=600, height =800 , bg='grey')
                  saveinginfo = Label(saving, text = "Save with same title to overwrite or change title to save as new plan",
                                      bg = 'steelblue')
                  saveinginfo.grid(row = 0, column = 0 , columnspan = 4)
                  self.entrybox = Entry (saving , width = 52, fg = 'white' , bg = 'black' , borderwidth = 10)
                  self.entrybox.grid(row=1,column=0,columnspan=4)
                  #This line takes the full address of th efile chosen and only uses the actual file name
                  #It then inserts the filename into the entry box to let the user easily overwrite an existing floor plan
                  self.entrybox.insert(0 , os.path.basename(self.filename))

                  submitbtn = Button (saving , text = "Save to File" , command= saveandexit , bg = 'steelblue')
                  submitbtn.grid(row=2,column=0,columnspan=4)
              

              seatingcolourdict = {} #Dictionary holding table and corresponding colour
              tabledict = {} #Dictionary holding all the tables and their functions/information

              import glob

              if len(glob.glob("*.txt")) == 0:

                     def returntofloormenu():
                            self.master.CreateFunctions(FloorPlanMenu,FloorPlanMenu)
                            self.master.show_frame(FloorPlanMenu)
                            noPlans.destroy()
                            

                     noPlans = Toplevel(width=800, height =800 , bg='grey')
                     noFloors = Label(noPlans, text = "No saved floor plans. Create one, then return if needed.",
                                         bg = 'steelblue',width = 70, font=("Courier New", 10))
                     noFloors.grid(row = 0, column = 0 , columnspan = 4)

                     returntomenu = Button (noPlans , text = "Return to Menu" , command= returntofloormenu, bg = 'steelblue')
                     returntomenu.grid(row=2,column=0,columnspan=4)

              else:
 
                     from tkinter.filedialog import askopenfilename #Allows us to use tkinter file management
                     # show an "Open" dialog box and return the path to the selected file
                     self.filename = askopenfilename(filetypes = [("Text Files", ".txt")]) 
                     
                     #Opens the chosen floor plan file and reads the colour of each table and saves it to a list
                     textfile = open(self.filename,"r") 
                     with open(self.filename) as f:
                         seatingcolourlist = f.read().splitlines()
                     
                     count1 = 1 # This iterates through the list and converts it to a dictionary
                     for x in seatingcolourlist :
                            seatingcolourdict[str(count1)] = x
                            count1+=1

                     info = Label (self, text = "Click a table to edit it ", font = "Verdana 15", padx = 10, pady = 10,
                                   bg = 'steelblue'  , fg = 'black' , relief = "raised" , borderwidth = 5)
                     info.grid(row=0,column=1, columnspan = 3) # Creates a label to inform the user

                     #Creates key of colours and corresponding seating capacity
                     key = Label (self, text = "Key: ", font = "Verdana 15" ,  bg = 'steelblue',
                                  fg = 'black' , relief = "raised" ).grid(row=1,column=0)
                     twoseater = Label (self, text = "2 Seats ", font = "Verdana 15" , bg = "yellow", padx = 5).grid(row=1,column=1)
                     fourseater = Label (self, text = "4 Seats  ", font = "Verdana 15" , bg = "orange", padx = 5).grid(row=1,column=2)
                     sixseater = Label (self, text = "6 Seats ", font = "Verdana 15" , bg = "red", padx = 5).grid(row=1,column=3)
                     tenseater = Label (self, text = "10 Seats ", font = "Verdana 15" , bg = "cyan", padx = 5).grid(row=1,column=4)
              
                     rowNumber = 2
                     columnNumber = 0

                     colourvalues = {"yellow" : "2" , "orange" : "4" , "red" : "6" , "cyan" : "10"}
                     #Creates the griid of tables, coloured according to the saved floorplan
                     for n in seatingcolourdict:
                            tabledict[n] = Button(self,text = "Table " + str(n), bg = seatingcolourdict[n],
                                                  font = "Verdana 12" ,width = 10, height = 5,command=lambda n=n: seatchoice(n))
                            tabledict[n].grid(row = rowNumber, column = columnNumber, padx = 5, pady = 10)

                            columnNumber += 1
                            if columnNumber == 5:
                                   columnNumber = 0
                                   rowNumber += 1

                     #Button at bottom of screen to open options to leave floor plan creation
                     returntofloormenu = Button(self, text = "Return to floor menu" , pady = 10 , padx = 10, font = "Verdana 12",
                                              command = leavecreationwarning ,  bg = 'steelblue'  , fg = 'black' , relief = "raised" , borderwidth = 5 )
                     returntofloormenu.grid(row = 7 , column = 0 , columnspan = 2)

                     #Button at bottom of screen to run nameenter function and then save and exit
                     saveandexitbtn = Button (self, text = "Save Floor Plan and Exit to Menu" , pady = 10 , padx = 10, font = "Verdana 12",
                                              command =nameenter ,  bg = 'steelblue'  , fg = 'black' , relief = "raised" , borderwidth = 5 )
                     saveandexitbtn.grid(row = 7 , column = 2 , columnspan = 3)
                     #Small fix to keep GUI looking clean
                     self.grid_columnconfigure(1, weight=1)
                     self.grid_columnconfigure(3, weight=1)

       ##              newvalue = {}
       ##              count2 = 1
       ##              for x in seatingcolourlist :
       ##                     z = colourvalues[x] #translates colour to seating capacity 
       ##                     newvalue[str(count2)] = z
       ##                     count2+=1
              

class CreateNewCustomer(Frame):
       def __init__(self, parent, controller):
              Frame.__init__(self, parent)
              Frame.config(self, bg= "grey")
              self.grid_columnconfigure(1, weight=1)

              #Creates a title label to tell the user what to do
              TitleText = Label(self, text="Please enter the customer's details into the correct input boxes",  bg = 'steelblue'
                                ,fg = 'black' , borderwidth = 5, pady = 10 , padx = 10, relief= 'raised', font=("Courier New", 10))
              TitleText.grid(row = 0 , column = 0 , columnspan = 3)

              #Creates the labels for each entry field
              name = Label(self, text="Name  :" , bg = 'grey', pady = 30 , padx = 10)
              name.grid(row = 1 , column = 0, sticky = W)

              phone = Label(self, text="Phone Number  :" , bg = 'grey', pady = 30 , padx = 10)
              phone.grid(row = 2 , column = 0, sticky = W)

              email = Label(self, text="Email Address  :" , bg = 'grey', pady = 30 , padx = 10)
              email.grid(row = 3 , column = 0, sticky = W)

              self.customerdetailsdict = {}
              
              #creates the entry fields for the details
              for x in range (1,4):
                     self.customerdetailsdict[x] = Entry(self , width = 50, fg = 'white' , bg = 'black' , borderwidth = 10 )
                     self.customerdetailsdict[x].grid(row= x, column = 1, columnspan = 2)
                          
              #Gives the radiobuttons an interger data type
              r = IntVar()
              r.set(1)

              #Creates the choice buttons to give the user the ability to opt in to recieve promotional emails
              OptIn = Label(self, text="Opt In to Promotional Emails :" , bg = 'grey', pady = 30 , padx = 10)
              OptIn.grid(row = 5 , column = 0, sticky = W)
              Radiobutton(self, text = "Yes" , variable = r , value = 1 , bg = 'grey').grid(row= 5, column = 1)
              Radiobutton(self, text = "No" , variable = r , value = 2 ,  bg = 'grey').grid(sticky = E,row= 5, column = 2)

              #A selection of premade tags refering to common food requests as well as the choice to create a custom tag
              global TAGS
              TAGS = [
                     "No Tag Needed",
                     "VIP",
                     "Vegetarian",
                     "Vegan",
                     "Gluten Free",
                     "Kosher",
                     "Halal",
                     "Nut Free",
                     "Other Tag Needed"
                     ] 

              #A function used to create extra fields for the user to enter custom tags if needed and a list of selected tags
              def CheckCustom(chosen):
                     
                     self.chosentaglist.append(chosen) #Creates list of tags used
                     
                     #Creates custom tag entry fields
                     if chosen == "Other Tag Needed":
                            
                            self.Tags = Label(self, text="Custom Tag Choice : " ,bg = 'grey',pady = 30,padx = 10)
                            self.Inputdict[self.count] = Entry(self , width = 50, fg = 'white' , bg = 'black' , borderwidth = 10 )
                            
                            self.Tags.grid(row = self.rowY+10, column = 0, sticky = W)
                            self.Inputdict[self.count].grid(row= self.rowY+10, column = 1, columnspan = 2)
                            self.rowY +=1
                            self.count += 1
                     #Deletes custom entry fields if not needed
                     else:
                            if self.count == 0:
                                   pass
                            else:
                                   self.Tags.grid_forget()
                                   self.Inputdict[self.count-1].grid_forget()

              #Activated when save button pressed. Opens the customer database and enters information
              def savedetails():
                     import openpyxl #Excel file manager module

                     #Opens excel file and sheet containg database
                     theFile = openpyxl.load_workbook('Customer Details Form.xlsx')
                     currentSheet = theFile['Sheet1']

                     X = 1
                     Y = currentSheet.max_row+1 #Selects the next empty row
                     
                     #Inputs name, number and email into database
                     for n in self.customerdetailsdict:
                            enteredvalue = self.customerdetailsdict[n].get()
                            print(enteredvalue)

                            currentSheet.cell(column= X, row=Y, value= str(enteredvalue))
                            X += 1

                     #Inputs opt in or out into database
                     currentSheet.cell(column= 4, row=Y, value= r.get())
                     

                     #Inputs all slected tags into database if no custom tags were selected
                     if self.count== 0:
                            for x in range (0,3):
                                   currentSheet.cell(column= (5 + int(x)), row=Y, value= str(self.chosentaglist[x]))

                     #Inputs exsting tags then custom tags
                     else:
                            y = 3
                            y = y - self.count

                            for x in range(0,y):
                                   currentSheet.cell(column= (5 + int(x)), row=Y, value= str(self.chosentaglist[x]))

                            for x in range (0, self.count):
                                   customtag = self.Inputdict[x].get()
                                   currentSheet.cell(column= (5+ y + int(x)), row=Y, value= str(customtag))
                     #Saves the excel file               
                     theFile.save('Customer Details Form.xlsx')

              def exit():

                     self.exitwindow = Toplevel(width=500, height =800 , bg='grey')

                     TitleText = Label(self.exitwindow, text="Are you sure you want to exit the customer details creator? \n Unsaved work will be deleted",
                                       bg = 'steelblue',fg = 'black' , borderwidth = 5, pady = 10 , relief= 'raised', font=("Courier New", 10))
                     TitleText.pack()

                     YesBtn = Button(self.exitwindow,  text = "Yes - Exit to Customer Details Menu ",  bg = 'steelblue', fg = 'black',
                                     pady = 10, padx = 10 , relief= 'raised', font=("Courier New", 10), command = ExittoCustomer)
                     YesBtn.pack()

                     NoBtn = Button(self.exitwindow,  text = "No - Return to Creation ",  bg = 'steelblue', fg = 'black',
                                     pady = 10 , relief= 'raised', font=("Courier New", 10) , command = ReturnToCreation)
                     NoBtn.pack()

              def ExittoCustomer():
                     self.master.CreateFunctions(CustomerDetails,CustomerDetails)
                     self.exitwindow.destroy()

              def ReturnToCreation():
                     self.exitwindow.destroy()

                     
              #Constants and dictionaries/lists
              TagMenudict = {}
              self.Inputdict = {}
              self.chosentaglist = []
              self.count = 0
              self.rowY=0
              
              #Creates the menus for the user to choose tags           
              for z in range (0,3):

                     Tags = Label(self, text="Tags Choice " + str(z + 1) + ":" , bg = 'grey', pady = 30 , padx = 10)
                     Tags.grid(row = (z+6), column = 0, sticky = W)

                     variable = StringVar()

                     TagMenudict[x] = OptionMenu(self, variable, *TAGS, command = CheckCustom)
                     TagMenudict[x].grid(row = (6 + z) , column = 1 , columnspan = 2)

              #Save customer details button
              savecustomer = Button(self, text = "Save customer details", bg = 'steelblue' , fg = 'black', borderwidth = 5,
                                    pady = 10 , padx = 10, relief= 'raised', command = savedetails ,font=("Courier New", 10))
              savecustomer.grid(row = 15, column = 1, sticky = W)

              #A exit button to leave to menu
              exit = Button(self, text = "Exit Button",  bg = 'steelblue'  , fg = 'black',borderwidth = 5, command = exit
                            , pady = 10 , padx = 10, relief= 'raised', font=("Courier New", 10))
              exit.grid(row = 15, column = 1, sticky = E)

class CustomerDetailsEditor(Frame):
       def __init__(self, parent, controller):
              Frame.__init__(self, parent)
              Frame.config(self, bg= "grey")
              self.grid_columnconfigure(1, weight=1)

              self.Inputdict = {}
              self.TagMenudict = {}
              self.variabledict = {}
              self.selectedTags = {}

              self.TagCount = 3
              self.count = 0
              self.rowY=0

              self.chosentaglist = []
              CUSTOMERS = []


              def InputCustomer(customerName):
                     #Stores a selected customers details to be inputted into the form
                     self.selectedcustomer = []                     

                     import openpyxl #Excel file manager module
                     #Opens excel file and sheet containg database and selects the fist column (customer names)
                     theFile = openpyxl.load_workbook('Customer Details Form.xlsx')
                     currentSheet = theFile['Sheet1']

                     #Iterates through the name column to find the row containing the selected customer
                     for i in range(1,currentSheet.max_row+1):
                            if currentSheet.cell(row=i, column=1).value == customerName:
                                   self.customerdetailrow = i
                                   #Inputs selected custoemrs setails into the self.selectedcustomer list
                                   for j in range(1, currentSheet.max_column+1):
                                          self.selectedcustomer.append(currentSheet.cell(row=i, column=j).value)

                     #Iterates through the entry fields and inputs the correct data from the selectedcustomer list
                     for x in range (1,4):
                            self.customerdetailsdict[x].delete(0, END)
                            self.customerdetailsdict[x].insert(0,self.selectedcustomer[x-1])

                     #Selects the button for whether or not they opt in or out to promo emails
                     if self.selectedcustomer[3] == 1:
                            optin.select()
                     else:
                            optout.select()

                     for x in range(0,3):
                            
                            self.TagMenudict[x].config(state = "normal")
                            if self.selectedcustomer[4+x]in TAGS:
                                          self.variabledict[x].set(self.selectedcustomer[4+x])
                                          CheckCustom(self.selectedcustomer[4+x])
                            else:
                                   self.variabledict[x].set("Other Tag Needed")
                                   CheckCustom("Other Tag Needed")

              #A function used to create extra fields for the user to enter custom tags if needed and a list of selected tags
              def CheckCustom(chosen):

                     #Creates a dictionary of chosen tags to be easily saved into the database
                     remainder = self.TagCount % 3
                     self.selectedTags[remainder] = chosen    
                     print(self.selectedTags)
                     
                     #Creates custom tag entry fields
                     if chosen == "Other Tag Needed":
                            
                            self.Tags = Label(self, text="Custom Tag Choice : " ,bg = 'grey',pady = 30,padx = 10)
                            self.Inputdict[self.count] = Entry(self , width = 50, fg = 'white' , bg = 'black' , borderwidth = 10 )
                            
                            self.Tags.grid(row = self.rowY+10, column = 0, sticky = W)
                            self.Inputdict[self.count].grid(row= self.rowY+10, column = 1, columnspan = 2)

                            self.rowY +=1
                            self.count += 1
                            
                            
                     #Deletes custom entry fields if not needed
                     else:
                            if self.count == 0:
                                   pass
                            else:
                                   self.Tags.grid_forget()
                                   self.Inputdict[self.count - 1].grid_forget()

                     self.TagCount += 1

              #Activated when save button pressed. Opens the customer database and enters information
              def savedetails():
                     import openpyxl #Excel file manager module

                     print(self.chosentaglist)

                     #Opens excel file and sheet containg database
                     theFile = openpyxl.load_workbook('Customer Details Form.xlsx')
                     currentSheet = theFile['Sheet1']

                     X = 1
                     Y = self.customerdetailrow #Selects the next empty row
                     
                     #Inputs name, number and email into database
                     for n in self.customerdetailsdict:
                            enteredvalue = self.customerdetailsdict[n].get()
                            print(enteredvalue)

                            currentSheet.cell(column= X, row=Y, value= str(enteredvalue))
                            X += 1

                     #Inputs opt in or out into database
                     currentSheet.cell(column= 4, row=Y, value= r.get())
                     
                     for x in range(0,3):
                            currentSheet.cell(column= (5+x), row=Y, value= str(self.selectedTags[x]))

                     #Saves the excel file               
                     theFile.save('Customer Details Form.xlsx')


##                     #Inputs all slected tags into database if no custom tags were selected
##                     taglistlength = len(self.chosentaglist) - 1
##                     if self.count== 0:
##                            for x in range (0,3):
##                                   currentSheet.cell(column= (7-x), row=Y, value= str(self.chosentaglist[taglistlength-x]))
##
##                     #Inputs exsting tags then custom tags
##                     else:
##                            y = 3
##                            y = y - self.count
##
##                            for x in range(0,y):
##                                   currentSheet.cell(column= (5 + int(x)), row=Y, value= str(self.chosentaglist[x]))
##
##                            for x in range (0, self.count):
##                                   customtag = self.Inputdict[x].get()
##                                   currentSheet.cell(column= (5+ y + int(x)), row=Y, value= str(customtag))


              #Creates a title label to tell the user what to do
              TitleText = Label(self, text="Please select a customer. Their details can then be edited.",  bg = 'steelblue'
                                ,fg = 'black' , borderwidth = 5, pady = 10 , padx = 10, relief= 'raised', font=("Courier New", 10))
              TitleText.grid(row = 0 , column = 0 , columnspan = 3, pady = (5,10))


              import openpyxl #Excel file manager module

              #Opens excel file and sheet containg database and selects the fist column (customer names)
              theFile = openpyxl.load_workbook('Customer Details Form.xlsx')
              currentSheet = theFile['Sheet1']
              first_column = currentSheet['A']

              #Program iterates through names and adds them to a list to be accessed by the menu
              for x in range(len(first_column)):
                     if currentSheet.cell(row=x+1, column=1).value == None:
                            pass
                     else:
                            CUSTOMERS.append(first_column[x].value) 

              #Removes the header of the column (Names) from the choices
              CUSTOMERS.remove("Name")

              #Creates the options menu with all customer names in the list
              variable = StringVar()
              SelectCustomer = OptionMenu(self, variable, *CUSTOMERS, command = InputCustomer)
              SelectCustomer.grid(row = 1 , column = 1 , columnspan = 2)
              
              #Creates the labels for each entry field
              name = Label(self, text="Name  :" , bg = 'grey', pady = 30 , padx = 10)
              name.grid(row = 2, column = 0, sticky = W)

              phone = Label(self, text="Phone Number  :" , bg = 'grey', pady = 30 , padx = 10)
              phone.grid(row = 3 , column = 0, sticky = W)

              email = Label(self, text="Email Address  :" , bg = 'grey', pady = 30 , padx = 10)
              email.grid(row = 4 , column = 0, sticky = W)

              
              #creates the entry fields for the details like before
              self.customerdetailsdict = {}
              for x in range (1,4):
                     self.customerdetailsdict[x] = Entry(self , width = 50, fg = 'white' , bg = 'black' , borderwidth = 10 )
                     self.customerdetailsdict[x].grid(row= x+1, column = 1, columnspan = 2)

              #Gives the radiobuttons an interger data type
              r = IntVar()
              r.set(1)

              #Creates the choice buttons to give the user the ability to opt in to recieve promotional emails
              OptIn = Label(self, text="Opt In to Promotional Emails :" , bg = 'grey', pady = 30 , padx = 10)
              OptIn.grid(row = 5 , column = 0, sticky = W)
              optin = Radiobutton(self, text = "Yes" , variable = r , value = 1 , bg = 'grey')
              optin.grid(row= 5, column = 1)
              optout = Radiobutton(self, text = "No" , variable = r , value = 2 ,  bg = 'grey')
              optout.grid(sticky = E, row= 5, column = 1)

              
              for z in range (0,3):

                     Tags = Label(self, text="Tags Choice " + str(z + 1) + ":" , bg = 'grey', pady = 30 , padx = 10)
                     Tags.grid(row = (z+6), column = 0, sticky = W)

                     self.variabledict[z] = StringVar()

                     self.TagMenudict[z] = OptionMenu(self, self.variabledict[z], *TAGS, command = CheckCustom)
                     self.TagMenudict[z].grid(row = (6 + z) , column = 1 , columnspan = 2)

                     self.TagMenudict[z].config(state = "disabled")
              
              #Save customer details button
              savecustomer = Button(self, text = "Save customer details", bg = 'steelblue' , fg = 'black', borderwidth = 5,
                                    pady = 10 , padx = 10, relief= 'raised' ,font=("Courier New", 10), command = savedetails)
              savecustomer.grid(row = 15, column = 1, sticky = W)

              #A exit button to leave to menu
              exit = Button(self, text = "Exit Button",  bg = 'steelblue'  , fg = 'black',borderwidth = 5,
                            pady = 10 , padx = 10, relief= 'raised', font=("Courier New", 10))
              exit.grid(row = 15, column = 1, sticky = E)

#Exit the program, giving a quick confirmation warning
class ExitProgram(Frame):
       def __init__(self, parent, controller):
              Frame.__init__(self, parent)
              Frame.config(self, bg= "grey")

              TitleText = Label(self, text="Are you sure you want to exit the program?",  bg = 'steelblue',
                                fg = 'black' , borderwidth = 5, pady = 10 , relief= 'raised', font=("Courier New", 10))
              TitleText.pack()

              YesBtn = Button(self,  text = "Yes - Exit Program ",  bg = 'steelblue', fg = 'black',
                              pady = 10, padx = 10 , relief= 'raised', font=("Courier New", 10), command = self.ExitOut)
              YesBtn.pack()

              NoBtn = Button(self,  text = "No - Return to Menu ",  bg = 'steelblue', fg = 'black',
                              pady = 10 , relief= 'raised', font=("Courier New", 10) , command = self.ReturnToMenu)
              NoBtn.pack()

       def ExitOut(self):
              exit(0)

       def ReturnToMenu(self):
              self.master.show_frame(MainMenu)


global tableSeatedCustomer
tableSeatedCustomer = {}

global tableReservedCustomer
tableReservedCustomer = {}        

#Runs the Program
app = RestaurantManager()
app.mainloop()


